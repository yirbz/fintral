import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.services.pipeline.base import BaseProcessor, ProcessingResult
from app.services.pipeline.classifier import XLSX_EXTENSIONS


REQUIRED_COLUMNS = [
    "★ RNC Proveedor",
    "★ Razón Social",
    "★ NCF",
    "★ Fecha Factura",
    "★ Monto Total",
]

COLUMN_MAPPING = {
    "★ RNC Proveedor": "RNC Proveedor",
    "★ Razón Social": "Razón Social",
    "★ NCF": "NCF",
    "★ Fecha Factura": "Fecha Factura",
    "Fecha Pago": "Fecha Pago",
    "★ Monto Total": "Monto Total",
    "ITBIS": "ITBIS",
    "Tipo B/S (606)": "Tipo B/S",
    "Forma de Pago": "Forma Pago",
    "Moneda": "Moneda",
    "Tipo Transacción": "Tipo Transacción",
    "Categoría": "Categoría",
    "Descripción": "Descripción",
}
VALID_GOODS_SERVICES = {f"{i:02d}" for i in range(1, 12)}
VALID_PAYMENT_METHODS = {"1", "2", "3", "4", "5", "6", "7"}
VALID_CURRENCIES = {"DOP", "USD", "EUR"}
VALID_TRANSACTION_TYPES = {"income", "expense"}


class XLSXProcessor(BaseProcessor):
    """Processor for XLSX bulk invoice imports."""

    name = "xlsx_processor"

    def can_process(self, file_path: str, file_type: str) -> bool:
        ext = file_path.split(".")[-1].lower().replace(".", "")
        return ext in [e.replace(".", "") for e in XLSX_EXTENSIONS]

    def process(self, file_path: str, **kwargs) -> ProcessingResult:
        errors = []
        invoices = []

        try:
            wb = load_workbook(file_path, data_only=True)
            
            # Find Facturas sheet
            ws = None
            for sheet_name in wb.sheetnames:
                if "Facturas" in sheet_name:
                    ws = wb[sheet_name]
                    break
            
            if not ws:
                return ProcessingResult(
                    success=False,
                    error="No se encontró hoja 'Facturas'",
                    source_type="xlsx",
                    confidence=0.0,
                )

            header_row = self._find_header_row(ws)
            if not header_row:
                return ProcessingResult(
                    success=False,
                    error="No se encontró fila de encabezados",
                    source_type="xlsx",
                    confidence=0.0,
                )

            headers = self._parse_headers(ws, header_row)
            validation_errors = self._validate_headers(headers)
            if validation_errors:
                return ProcessingResult(
                    success=False,
                    error=f"Encabezados inválidos: {', '.join(validation_errors)}",
                    source_type="xlsx",
                    confidence=0.0,
                )

            row_num = header_row + 1
            while row_num <= ws.max_row:
                row_data = self._parse_row(ws, row_num, headers)
                if self._is_empty_row(row_data):
                    row_num += 1
                    continue

                row_errors = self._validate_row(row_data, row_num)
                if row_errors:
                    errors.extend(row_errors)
                else:
                    invoices.append(self._normalize_row(row_data))

                row_num += 1

            if not invoices:
                return ProcessingResult(
                    success=False,
                    error="No se encontraron facturas válidas",
                    source_type="xlsx",
                    confidence=0.0,
                )

            return ProcessingResult(
                success=True,
                data={"invoices": invoices, "errors": errors},
                source_type="xlsx",
                confidence=1.0 if not errors else 0.7,
                warnings=errors,
            )

        except Exception as exc:
            return ProcessingResult(
                success=False,
                error=f"Error procesando XLSX: {str(exc)}",
                source_type="xlsx",
                confidence=0.0,
            )

    def _find_header_row(self, ws) -> int:
        # Template has headers in row 1
        return 1

    def _parse_headers(self, ws, row_num: int) -> Dict[str, int]:
        headers = {}
        for col_num in range(1, ws.max_column + 1):
            value = ws.cell(row_num, col_num).value
            if value:
                headers[str(value).strip()] = col_num
        return headers

    def _validate_headers(self, headers: Dict[str, int]) -> List[str]:
        missing = [col for col in REQUIRED_COLUMNS if col not in headers]
        return [f"Falta columna: {col}" for col in missing]

    def _parse_row(self, ws, row_num: int, headers: Dict[str, int]) -> Dict[str, Any]:
        row_data = {}
        for col_name, col_num in headers.items():
            value = ws.cell(row_num, col_num).value
            row_data[col_name] = value
        return row_data

    def _is_empty_row(self, row_data: Dict[str, Any]) -> bool:
        required_fields = ["★ RNC Proveedor", "★ NCF", "★ Monto Total"]
        return not any(row_data.get(f) for f in required_fields)

    def _validate_row(self, row_data: Dict[str, Any], row_num: int) -> List[str]:
        errors = []

        rnc = str(row_data.get("★ RNC Proveedor", "")).strip()
        if not re.match(r"^\d{9}$|^\d{3}-\d{7}-\d{1}$", rnc):
            errors.append(f"Fila {row_num}: RNC inválido")

        ncf = str(row_data.get("★ NCF", "")).strip().upper()
        if not re.match(r"^[BE]\d{2}\d{8,10}$", ncf):
            errors.append(f"Fila {row_num}: NCF inválido")

        total = row_data.get("★ Monto Total")
        try:
            if float(total) <= 0:
                errors.append(f"Fila {row_num}: Monto debe ser mayor a 0")
        except (TypeError, ValueError):
            errors.append(f"Fila {row_num}: Monto inválido")

        tipo_bs = str(row_data.get("Tipo B/S (606)", "")).strip()
        if tipo_bs and tipo_bs not in VALID_GOODS_SERVICES:
            errors.append(f"Fila {row_num}: Tipo B/S inválido (debe ser 01-11)")

        forma_pago = str(row_data.get("Forma de Pago", "")).strip()
        if forma_pago and forma_pago not in VALID_PAYMENT_METHODS:
            errors.append(f"Fila {row_num}: Forma pago inválido (debe ser 1-7)")

        return errors

    def _normalize_row(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "vendor_tax_id": str(row_data.get("★ RNC Proveedor", "")).strip(),
            "vendor_name": str(row_data.get("★ Razón Social", "")).strip(),
            "invoice_number": str(row_data.get("★ NCF", "")).strip().upper(),
            "invoice_date": self._parse_date(row_data.get("★ Fecha Factura")),
            "payment_date": self._parse_date(row_data.get("Fecha Pago")),
            "total_amount": float(row_data.get("★ Monto Total", 0)),
            "tax_amount": self._parse_number(row_data.get("ITBIS")),
            "goods_services_type": str(row_data.get("Tipo B/S (606)", "")).strip() or None,
            "payment_method": str(row_data.get("Forma de Pago", "")).strip() or None,
            "currency": self._parse_currency(row_data.get("Moneda")),
            "transaction_type": self._parse_transaction_type(row_data.get("Tipo Transacción")),
            "category": str(row_data.get("Categoría", "")).strip() or "sin_categoria",
            "description": str(row_data.get("Descripción", "")).strip(),
        }

    def _parse_date(self, value: Any) -> Optional[str]:
        if not value:
            return None
        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d")
            dt = datetime.strptime(str(value), "%Y-%m-%d")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return None

    def _parse_number(self, value: Any) -> Optional[float]:
        if not value:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _parse_currency(self, value: Any) -> str:
        currency = str(value).upper().strip() if value else "DOP"
        return currency if currency in VALID_CURRENCIES else "DOP"

    def _parse_transaction_type(self, value: Any) -> str:
        vt = str(value).lower().strip() if value else "expense"
        return vt if vt in VALID_TRANSACTION_TYPES else "expense"

    def create_template_bytes(self) -> bytes:
        wb = create_invoice_template()
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


def create_invoice_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Instrucciones"

    ws.append(["INSTRUCCIONES PARA IMPORTAR FACTURAS"])
    ws.append([])
    ws.append(["1. Complete los datos en la hoja 'Facturas'"])
    ws.append(["2. Los campos en negrita son obligatorios"])
    ws.append(["3. No agregue filas vacías entre datos"])
    ws.append(["4. Use el formato de fecha YYYY-MM-DD"])
    ws.append(["5. Guarde el archivo como .xlsx"])
    ws.append([])

    instructions_data = [
        ["CAMPO", "DESCRIPCIÓN", "FORMATO", "EJEMPLO"],
        ["RNC Proveedor", "Registro Nacional del Contribuyente (9 dígitos)", "Texto", "123456789"],
        ["Razón Social", "Nombre oficial del proveedor", "Texto", "Empresa SRL"],
        ["NCF", "Número de Comprobante Fiscal (B/E + dígitos)", "Texto", "B0100000001"],
        ["Fecha Factura", "Fecha de emisión (YYYY-MM-DD)", "Fecha", "2026-01-15"],
        ["Fecha Pago", "Fecha de pago (YYYY-MM-DD)", "Fecha", "2026-02-15"],
        ["Monto Total", "Monto total con impuestos", "Número", "1000.00"],
        ["ITBIS", "Monto ITBIS (18%)", "Número", "150.00"],
        ["Tipo B/S", "Tipo DGII 606 (01-11)", "Texto", "02"],
        ["Forma Pago", "Forma de pago (1-7)", "Texto", "1"],
        ["Moneda", "Moneda (DOP/USD/EUR)", "Texto", "DOP"],
        ["Tipo Transacción", "income/expense", "Texto", "expense"],
        ["Categoría", "Categoría de gasto", "Texto", "servicios"],
        ["Descripción", "Descripción adicional", "Texto", "Servicios de consultoría"],
    ]

    for row in instructions_data:
        ws.append(row)

    ws_facturas = wb.create_sheet("Facturas")

    headers = [
        ("RNC Proveedor", True),
        ("Razón Social", True),
        ("NCF", True),
        ("Fecha Factura", True),
        ("Fecha Pago", False),
        ("Monto Total", True),
        ("ITBIS", False),
        ("Tipo B/S", False),
        ("Forma Pago", False),
        ("Moneda", False),
        ("Tipo Transacción", False),
        ("Categoría", False),
        ("Descripción", False),
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, (header, _) in enumerate(headers, 1):
        cell = ws_facturas.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws_facturas.append(["501201234", "Servicios Técnicos SRL", "B0100001001", "2026-01-01", "2026-01-31", 5000.00, 750.00, "02", "2", "DOP", "expense", "servicios", "Mantenimiento sistema"])

    col_widths = [15, 25, 18, 15, 15, 12, 12, 10, 10, 8, 15, 15, 30]
    for idx, width in enumerate(col_widths, 1):
        ws_facturas.column_dimensions[get_column_letter(idx)].width = width

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50

    return wb


xlsx_processor = XLSXProcessor()