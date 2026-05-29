import csv
import io
import json
import os
import zipfile
from datetime import datetime, timedelta
from typing import List, Optional, Dict
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from app.models import Invoice


FISCAL_HEADERS = [
    "RNC",
    "Proveedor",
    "NCF",
    "Fecha",
    "Tipo",
    "Categoría",
    "Tipo Bien/Serv",
    "Descripción",
    "Base Imponible",
    "ITBIS",
    "Total",
    "Moneda",
    "Estado",
]


def _fiscal_row(inv: Invoice) -> list:
    base = (inv.total_amount or 0) - (inv.tax_amount or 0)
    return [
        inv.vendor_tax_id or "",
        inv.vendor_name or "",
        inv.invoice_number or "",
        inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
        {"income": "Ingreso", "expense": "Gasto"}.get(inv.transaction_type or "", inv.transaction_type or ""),
        inv.category or "",
        inv.goods_services_type or "",
        inv.description or "",
        f"{base:.2f}",
        f"{inv.tax_amount or 0:.2f}",
        f"{inv.total_amount or 0:.2f}",
        inv.currency or "DOP",
        "Procesado" if inv.processed else "Pendiente",
    ]

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "plantilla_excel")
_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'


# ── Template filler: modifies .xlsx at the XML level ────────────
# Preserves 100% of styles, merged cells, column widths, formulas.

class TemplateFiller:
    """Load an .xlsx template, fill cell values via raw XML, return bytes."""

    def __init__(self):
        self._files: Dict[str, bytes] = {}
        self._sheet_tree: etree._ElementTree = None
        self._sheet_root: etree._Element = None
        self._sheet_path: str = ''
        self._sheet_name: str = ''
        self._strings: List[str] = []
        self._ss_tree: etree._ElementTree = None
        self._ss_root: etree._Element = None

    def load(self, filename: str):
        path = os.path.join(TEMPLATE_DIR, filename)
        if not os.path.exists(path):
            raise FileNotFoundError(f"Plantilla no encontrada: {filename}")
        with zipfile.ZipFile(path) as z:
            self._files = {n: z.read(n) for n in z.namelist()}

        # Determine which sheet is NOT named 'Utilitario'
        wb = etree.fromstring(self._files['xl/workbook.xml'])
        wns = {'s': _NS, 'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
        sheet_target = None
        for sheet in wb.findall('.//s:sheet', wns):
            name = sheet.get('name', '')
            if 'utilitario' not in name.lower():
                rid = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                # Find target from relationships
                rels = etree.fromstring(self._files['xl/_rels/workbook.xml.rels'])
                for rel in rels:
                    if rel.get('Id') == rid:
                        sheet_target = rel.get('Target', '')
                        break
                self._sheet_name = name
                break

        if not sheet_target:
            raise RuntimeError('No data sheet found in template')

        self._sheet_path = f'xl/{sheet_target}'
        self._sheet_tree = etree.fromstring(self._files[self._sheet_path])
        self._sheet_root = self._sheet_tree

        # Parse shared strings
        if 'xl/sharedStrings.xml' in self._files:
            self._ss_tree = etree.fromstring(self._files['xl/sharedStrings.xml'])
            self._ss_root = self._ss_tree
            for si in self._ss_root.findall(f'{{{_NS}}}si'):
                t = si.find(f'{{{_NS}}}t')
                if t is not None:
                    self._strings.append(t.text or '')
                else:
                    texts = si.findall(f'.//{{{_NS}}}t')
                    self._strings.append(''.join(t2.text or '' for t2 in texts))

    @property
    def sheet_name(self) -> str:
        return self._sheet_name

    def header_map(self) -> Dict[str, int]:
        """Read headers from row 11 (1-indexed) → {name: 1-indexed col}.

        Keys are stripped of leading/trailing whitespace so template cells
        like 'Fecha Pago   ' match the dict key 'Fecha Pago'.
        Returns 1-indexed column numbers (consistent with write_cell).
        """
        mapping = {}
        sheet_data = self._sheet_root.find(f'{{{_NS}}}sheetData')
        for row in sheet_data.findall(f'{{{_NS}}}row'):
            if int(row.get('r')) == 11:
                for cell in row.findall(f'{{{_NS}}}c'):
                    ref = cell.get('r', '')
                    col_letter = ''.join(c for c in ref if c.isalpha())
                    col_1idx = self._col_letter_to_idx(col_letter) + 1  # 1-indexed
                    val = self._cell_value(cell).strip()
                    if val:
                        mapping[val] = col_1idx
                break
        return mapping

    def write_cell(self, row_1idx: int, col_1idx: int, value, cell_style: int = 0):
        """Write a value to cell (row, col) in 1-indexed notation.
        Preserves existing style if cell exists; uses cell_style for new cells."""
        ref = f'{self._col_idx_to_letter(col_1idx)}{row_1idx}'
        sheet_data = self._sheet_root.find(f'{{{_NS}}}sheetData')

        # Find or create row
        row_elem = None
        for row in sheet_data.findall(f'{{{_NS}}}row'):
            if int(row.get('r')) == row_1idx:
                row_elem = row
                break

        if row_elem is None:
            row_elem = etree.SubElement(sheet_data, f'{{{_NS}}}row')
            row_elem.set('r', str(row_1idx))

        # Find or create cell
        cell_elem = None
        for cell in row_elem.findall(f'{{{_NS}}}c'):
            if cell.get('r') == ref:
                cell_elem = cell
                break

        if cell_elem is None:
            cell_elem = etree.SubElement(row_elem, f'{{{_NS}}}c')
            cell_elem.set('r', ref)
            cell_elem.set('s', str(cell_style))

        # Write value
        self._remove_children(cell_elem, f'{{{_NS}}}v')
        self._remove_children(cell_elem, f'{{{_NS}}}is')
        v_elem = etree.SubElement(cell_elem, f'{{{_NS}}}v')

        if value is None or value == '':
            cell_elem.set('t', 'n')
            v_elem.text = ''
        elif isinstance(value, bool):
            cell_elem.set('t', 'b')
            v_elem.text = '1' if value else '0'
        elif isinstance(value, (int, float)):
            cell_elem.set('t', 'n')
            v_elem.text = str(value)
        else:
            cell_elem.set('t', 'str')
            v_elem.text = str(value)

    def clear_from(self, first_row_1idx: int):
        """Clear all data rows from first_row onward (single pass)."""
        sheet_data = self._sheet_root.find(f'{{{_NS}}}sheetData')
        for row in sheet_data.findall(f'{{{_NS}}}row'):
            if int(row.get('r')) >= first_row_1idx:
                for cell in row.findall(f'{{{_NS}}}c'):
                    v = cell.find(f'{{{_NS}}}v')
                    if v is not None:
                        v.text = ''
                        cell.set('t', 'n')

    def save(self) -> bytes:
        """Re-zip and return .xlsx bytes."""
        self._files[self._sheet_path] = etree.tostring(
            self._sheet_tree, xml_declaration=True, encoding='UTF-8', standalone=True)

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
            for name, data in self._files.items():
                z.writestr(name, data)
        buf.seek(0)
        return buf.getvalue()

    # ── Private helpers ──

    def _cell_value(self, cell) -> str:
        ct = cell.get('t', 'n')
        v = cell.find(f'{{{_NS}}}v')
        if v is None or not v.text:
            return ''
        if ct == 's':
            idx = int(v.text)
            raw = self._strings[idx] if 0 <= idx < len(self._strings) else ''
            return raw.strip()
        return v.text.strip()

    def _col_letter_to_idx(self, letters: str) -> int:
        idx = 0
        for ch in letters.upper():
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1

    def _col_idx_to_letter(self, idx_1idx: int) -> str:
        n = idx_1idx
        s = ''
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(ord('A') + r) + s
        return s

    def _remove_children(self, parent, tag: str):
        for child in list(parent):
            if child.tag == tag:
                parent.remove(child)


class ExportService:
    _SUBMISSION_REPORT_COLUMNS = {
        "606": [
            ("rnc_cedula", "RNC/Cédula"),
            ("tipo_id", "Tipo ID"),
            ("tipo_bienes_servicios", "Tipo B/S"),
            ("ncf", "NCF"),
            ("ncf_modificado", "NCF Modificado"),
            ("fecha_comprobante", "Fecha Comp."),
            ("fecha_pago", "Fecha Pago"),
            ("monto_servicios", "Monto Servicios"),
            ("monto_bienes", "Monto Bienes"),
            ("total_facturado", "Total Facturado"),
            ("itbis_facturado", "ITBIS Facturado"),
            ("itbis_retenido", "ITBIS Retenido"),
            ("itbis_proporcionalidad", "ITBIS Proporcionalidad"),
            ("itbis_costo", "ITBIS Costo"),
            ("itbis_adelantar", "ITBIS Adelantar"),
            ("itbis_percibido", "ITBIS Percibido"),
            ("tipo_retencion_isr", "Tipo Ret. ISR"),
            ("monto_retencion_renta", "Monto Ret. Renta"),
            ("isr_percibido", "ISR Percibido"),
            ("isc", "ISC"),
            ("otros_impuestos", "Otros Imp."),
            ("propina_legal", "Propina"),
            ("forma_pago", "Forma Pago"),
        ],
        "607": [
            ("rnc_comprador", "RNC/Cédula/Pasaporte"),
            ("tipo_id", "Tipo ID"),
            ("ncf", "NCF"),
            ("ncf_modificado", "NCF Modificado"),
            ("tipo_ingreso", "Tipo Ingreso"),
            ("fecha_comprobante", "Fecha Comp."),
            ("fecha_retencion", "Fecha Ret."),
            ("monto_facturado", "Monto Facturado"),
            ("itbis_facturado", "ITBIS Facturado"),
            ("itbis_retenido_terceros", "ITBIS Ret. Terceros"),
            ("itbis_percibido", "ITBIS Percibido"),
            ("retencion_renta_terceros", "Ret. Renta Terceros"),
            ("isr_percibido", "ISR Percibido"),
            ("isc", "ISC"),
            ("otros_impuestos", "Otros Imp."),
            ("propina_legal", "Propina"),
            ("efectivo", "Efectivo"),
            ("cheque_transferencia", "Cheque/Transf."),
            ("tarjeta", "Tarjeta"),
            ("credito", "Crédito"),
            ("bonos", "Bonos"),
            ("permuta", "Permuta"),
            ("otras_formas", "Otras Formas"),
        ],
        "608": [
            ("ncf", "NCF"),
            ("fecha_comprobante", "Fecha Comp."),
            ("tipo_anulacion", "Tipo Anulación"),
        ],
    }

    def submission_report_columns(self, format_code: str) -> List[Dict[str, str]]:
        return [
            {"key": key, "label": label}
            for key, label in self._SUBMISSION_REPORT_COLUMNS.get(format_code, [])
        ]

    def build_submission_report_row(
        self,
        format_code: str,
        invoice: Invoice,
        report_rnc: Optional[str] = None,
    ) -> Dict[str, str]:
        format_code = (format_code or "").replace("dgii_", "")
        if format_code == "606":
            fields = self._build_606_txt_fields(invoice, report_rnc)
        elif format_code == "607":
            fields = self._build_607_txt_fields(invoice)
        elif format_code == "608":
            fields = self._build_608_txt_fields(invoice)
        else:
            return {}

        keys = [key for key, _ in self._SUBMISSION_REPORT_COLUMNS.get(format_code, [])]
        return {key: fields[idx] if idx < len(fields) else "" for idx, key in enumerate(keys)}

    # ── 606 — Compras ────────────────────────────────────────────────────────

    def export_dgii_606(self, invoices: List[Invoice], report_rnc: Optional[str] = None, period: Optional[str] = None) -> bytes:
        """Exportación XLSX usando plantilla oficial DGII 606 vía XML puro."""
        tf = TemplateFiller()
        tf.load("606.xlsx")
        header_map = tf.header_map()

        rnc_value = self._only_digits(report_rnc) if report_rnc else ""
        period_value = period or self._derive_period(invoices) or ""
        tf.write_cell(4, 3, rnc_value)
        tf.write_cell(5, 3, period_value)
        tf.write_cell(6, 3, len(invoices))

        rows_data = []
        for row_idx, inv in enumerate(invoices, start=1):
            raw = self._parse_raw_data(inv.raw_extracted_data)
            goods_type = self._normalize_goods_type(inv.goods_services_type or raw.get("goods_services_type"))
            ncf = inv.invoice_number or raw.get("invoice_number") or ""
            vendor_tax_id = inv.vendor_tax_id or raw.get("vendor_tax_id")
            if self._ncf_document_type(ncf) in {"13", "17", "43", "47"}:
                vendor_tax_id = report_rnc
            rnc = self._only_digits(vendor_tax_id)
            tipo_id = self._tipo_id_from_tax_id(rnc)
            ncf_modified = raw.get("ncf_modified") or ""
            fecha_comprobante = self._format_date(inv.invoice_date or raw.get("invoice_date"))
            fecha_pago = self._format_date(raw.get("payment_date"))
            total = self._to_number(inv.total_amount) or self._to_number(raw.get("total_amount"))
            tax = self._to_number(inv.tax_amount) or self._to_number(raw.get("tax_amount"))
            base = self._compute_base(total, tax, raw)
            amount_services, amount_goods = self._split_base(base, inv, goods_type, raw)
            total_facturado = (amount_services or 0.0) + (amount_goods or 0.0)
            itbis_facturado = self._to_number(tax) or 0.0
            itbis_retenido = self._to_number(raw.get("itbis_retenido"))
            itbis_proporcionalidad = self._to_number(raw.get("itbis_proporcionalidad"))
            itbis_llevado_costo = self._to_number(raw.get("itbis_llevado_costo"))
            itbis_percibido = self._to_number(raw.get("itbis_percibido"))
            itbis_adelantar = itbis_facturado - (itbis_llevado_costo or 0.0)
            if itbis_adelantar < 0:
                itbis_adelantar = 0.0
            isr_retention_type = self._normalize_isr_retention(raw.get("isr_retention_type"))
            isr_retention_amount = self._to_number(raw.get("isr_retention_amount"))
            isr_percibido = self._to_number(raw.get("isr_percibido"))
            isc_amount = self._to_number(raw.get("isc_amount"))
            other_taxes = self._to_number(raw.get("other_taxes"))
            legal_tip = self._to_number(raw.get("legal_tip"))
            payment_method = self._normalize_payment_method(raw.get("payment_method"))
            status = self._build_606_status(rnc, ncf, fecha_comprobante, fecha_pago, total_facturado, itbis_facturado, itbis_retenido, isr_retention_type, isr_retention_amount)

            rows_data.append({
                "Líneas": row_idx,
                "RNC o Cédula": rnc or "",
                "Tipo Id": tipo_id or "",
                "Tipo Bienes y Servicios Comprados": goods_type or "",
                "NCF": ncf or "",
                "NCF ó Documento Modificado": ncf_modified or "",
                "Fecha Comprobante": fecha_comprobante or "",
                "Fecha Pago": fecha_pago or "",
                "Monto Facturado en Servicios": self._fmt_amount(amount_services, allow_zero=True),
                "Monto Facturado en Bienes": self._fmt_amount(amount_goods, allow_zero=True),
                "Total Monto Facturado": self._fmt_amount(total_facturado, allow_zero=True),
                "ITBIS Facturado": self._fmt_amount(itbis_facturado, allow_zero=True),
                "ITBIS Retenido": self._fmt_amount(itbis_retenido),
                "ITBIS sujeto a Proporcionalidad (Art. 349)": self._fmt_amount(itbis_proporcionalidad),
                "ITBIS llevado al Costo": self._fmt_amount(itbis_llevado_costo),
                "ITBIS por Adelantar": self._fmt_amount(itbis_adelantar, allow_zero=True),
                "ITBIS percibido en compras": self._fmt_amount(itbis_percibido),
                "Tipo de Retención en ISR": isr_retention_type or "",
                "Monto Retención Renta": self._fmt_amount(isr_retention_amount),
                "ISR Percibido en compras": self._fmt_amount(isr_percibido),
                "Impuesto Selectivo al Consumo": self._fmt_amount(isc_amount),
                "Otros Impuesto/Tasas": self._fmt_amount(other_taxes),
                "Monto Propina Legal": self._fmt_amount(legal_tip),
                "Forma de Pago": payment_method or "",
                "Estatus": status,
            })

        data_start = 12
        for i, data in enumerate(rows_data):
            row = data_start + i
            for name, value in data.items():
                col = header_map.get(name)
                if col is not None:
                    tf.write_cell(row, col, value)

        self._clear_remaining(tf, data_start, len(rows_data))
        return tf.save()

    # ── 607 — Ventas ─────────────────────────────────────────────────────────

    def export_dgii_607(self, invoices: List[Invoice], report_rnc: Optional[str] = None, period: Optional[str] = None, as_xls: bool = False) -> bytes:
        if as_xls:
            return self._export_607_xls(invoices, report_rnc, period)
        return self._export_607_csv(invoices, report_rnc, period)

    def _export_607_xls(self, invoices: List[Invoice], report_rnc: Optional[str] = None, period: Optional[str] = None) -> bytes:
        """Exportación XLSX usando plantilla oficial DGII 607 vía XML puro."""
        tf = TemplateFiller()
        tf.load("607.xlsx")
        header_map = tf.header_map()

        rnc_value = self._only_digits(report_rnc) if report_rnc else ""
        period_value = period or self._derive_period(invoices) or ""
        tf.write_cell(4, 3, rnc_value)
        tf.write_cell(5, 3, period_value)
        tf.write_cell(6, 3, len(invoices))

        rows_data = []
        for idx, inv in enumerate(invoices):
            raw = self._parse_raw_data(inv.raw_extracted_data)
            buyer_tax_id = inv.vendor_tax_id or raw.get("vendor_tax_id") or raw.get("buyer_tax_id")
            rnc = self._only_digits(buyer_tax_id)
            tipo_id = self._tipo_id_from_tax_id(rnc)
            ncf = inv.invoice_number or raw.get("invoice_number") or ""
            ncf_modified = raw.get("ncf_modified") or ""
            fecha = self._format_date(inv.invoice_date or raw.get("invoice_date"))
            fecha_retencion = self._format_date(raw.get("payment_date"))
            total = self._to_number(inv.total_amount) or self._to_number(raw.get("total_amount")) or 0.0
            tax = self._to_number(inv.tax_amount) or self._to_number(raw.get("tax_amount")) or 0.0
            amount_facturado = self._compute_base(total, tax, raw)
            itbis_retenido = self._to_number(raw.get("itbis_retenido")) or 0.0
            itbis_percibido = self._to_number(raw.get("itbis_percibido")) or 0.0

            issues = []
            if not ncf:
                issues.append("Falta NCF")
            if not fecha:
                issues.append("Falta fecha")
            if amount_facturado == 0:
                issues.append("Monto cero")
            status = "; ".join(issues) if issues else "OK"

            efectivo, cheque, tarjeta, credito, bonos, permuta, otras = self._payment_breakdown_607(
                raw.get("payment_method"),
                total,
            )

            rows_data.append({
                "No": idx + 1,
                "RNC/Cédula o Pasaporte": rnc or "",
                "Tipo Identificación": tipo_id or "",
                "Número Comprobante Fiscal": ncf,
                "Número Comprobante Fiscal Modificado": ncf_modified,
                "Tipo de Ingreso": "01",
                "Fecha Comprobante": fecha or "",
                "Fecha de Retención": fecha_retencion or "",
                "Monto Facturado": self._fmt_amount(amount_facturado, allow_zero=True),
                "ITBIS Facturado": self._fmt_amount(tax, allow_zero=True),
                "ITBIS Retenido por Terceros": self._fmt_amount(itbis_retenido),
                "ITBIS Percibido": self._fmt_amount(itbis_percibido),
                "Retención Renta por Terceros": "",
                "ISR Percibido": "",
                "Impuesto Selectivo al Consumo": "",
                "Otros Impuestos/Tasas": "",
                "Monto Propina Legal": "",
                "Efectivo": self._fmt_amount(efectivo),
                "Cheque/ Transferencia/ Depósito": self._fmt_amount(cheque),
                "Tarjeta Débito/Crédito": self._fmt_amount(tarjeta),
                "Venta a Crédito": self._fmt_amount(credito),
                "Bonos o Certificados de Regalo": self._fmt_amount(bonos),
                "Permuta": self._fmt_amount(permuta),
                "Otras Formas de Ventas": self._fmt_amount(otras),
                "Estatus": status,
            })

        data_start = 12
        for i, data in enumerate(rows_data):
            row = data_start + i
            for name, value in data.items():
                col = header_map.get(name)
                if col is not None:
                    tf.write_cell(row, col, value)

        self._clear_remaining(tf, data_start, len(rows_data))
        return tf.save()

    def _export_607_csv(self, invoices, report_rnc=None, period=None) -> bytes:
        """607 en CSV (fallback rápido)."""
        output = io.StringIO()
        writer = csv.writer(output)
        rnc_value = self._only_digits(report_rnc) if report_rnc else ""
        period_value = period or self._derive_period(invoices) or ""
        writer.writerow(["FORMULARIO 607 - COMPROBANTES DE VENTAS Y SERVICIOS"])
        writer.writerow([f"RNC Declarante: {rnc_value}", f"Período: {period_value}", f"Registros: {len(invoices)}"])
        writer.writerow([])
        writer.writerow(["RNC o Cédula", "Tipo ID", "NCF", "NCF Modificado", "Fecha Comprobante", "Fecha Pago",
                         "Monto Facturado", "ITBIS Facturado", "ITBIS Retenido por Terceros",
                         "ITBIS Percibido", "Forma de Pago", "Proveedor/Cliente", "Estatus 607"])
        for inv in invoices:
            raw = self._parse_raw_data(inv.raw_extracted_data)
            buyer_tax_id = inv.vendor_tax_id or raw.get("vendor_tax_id") or raw.get("buyer_tax_id")
            rnc = self._only_digits(buyer_tax_id)
            tipo_id = self._tipo_id_from_tax_id(rnc)
            ncf = inv.invoice_number or raw.get("invoice_number") or ""
            ncf_modified = raw.get("ncf_modified") or ""
            fecha_comprobante = self._format_date(inv.invoice_date or raw.get("invoice_date"))
            fecha_pago = self._format_date(raw.get("payment_date"))
            total = self._to_number(inv.total_amount) or self._to_number(raw.get("total_amount")) or 0.0
            tax = self._to_number(inv.tax_amount) or self._to_number(raw.get("tax_amount")) or 0.0
            amount_facturado = self._compute_base(total, tax, raw)
            itbis_retenido = self._to_number(raw.get("itbis_retenido")) or 0.0
            itbis_percibido = self._to_number(raw.get("itbis_percibido")) or 0.0
            payment_method = self._normalize_payment_method(raw.get("payment_method"))
            issues = []
            if not ncf:
                issues.append("Falta NCF")
            if not fecha_comprobante:
                issues.append("Falta fecha")
            if amount_facturado == 0:
                issues.append("Monto cero")
            status = "; ".join(issues) if issues else "OK"
            writer.writerow([rnc or "", tipo_id or "", ncf, ncf_modified, fecha_comprobante or "", fecha_pago or "",
                             self._fmt_amount(amount_facturado, allow_zero=True) or '',
                             self._fmt_amount(tax, allow_zero=True) or '',
                             self._fmt_amount(itbis_retenido) or '',
                             self._fmt_amount(itbis_percibido) or '',
                             payment_method or "", inv.vendor_name or "", status])
        return output.getvalue().encode("utf-8-sig")

    # ── 608 — Anulaciones ─────────────────────────────────────────────────────

    def export_dgii_608(self, invoices: List[Invoice], report_rnc: Optional[str] = None, period: Optional[str] = None, as_xls: bool = False) -> bytes:
        if as_xls:
            return self._export_608_xls(invoices, report_rnc, period)
        return self._export_608_csv(invoices, report_rnc, period)

    def _export_608_xls(self, invoices: List[Invoice], report_rnc: Optional[str] = None, period: Optional[str] = None) -> bytes:
        """Exportación XLSX usando plantilla oficial DGII 608 vía XML puro."""
        tf = TemplateFiller()
        tf.load("608.xlsx")
        header_map = tf.header_map()

        rnc_value = self._only_digits(report_rnc) if report_rnc else ""
        period_value = period or self._derive_period(invoices) or ""
        tf.write_cell(5, 3, rnc_value)
        tf.write_cell(6, 3, period_value)
        tf.write_cell(7, 3, len(invoices))

        rows_data = []
        for idx, inv in enumerate(invoices):
            raw = self._parse_raw_data(inv.raw_extracted_data)
            ncf = inv.invoice_number or raw.get("invoice_number") or ""
            fecha = self._format_date(inv.invoice_date or raw.get("invoice_date"))
            fecha_anulacion = ""
            if inv.cancelled_at:
                fecha_anulacion = inv.cancelled_at.strftime('%Y%m%d')
            elif inv.deleted_at:
                fecha_anulacion = inv.deleted_at.strftime('%Y%m%d')
            elif inv.updated_at:
                fecha_anulacion = inv.updated_at.strftime('%Y%m%d')
            tipo_anulacion = inv.cancellation_type or raw.get("cancellation_type") or "01"

            issues = []
            if not ncf:
                issues.append("Falta NCF")
            if not fecha_anulacion:
                issues.append("Falta fecha anulación")
            status = "; ".join(issues) if issues else "OK"

            rows_data.append({
                "Líneas": idx + 1,
                "Número de Comprobante Fiscal": ncf,
                "Fecha de Comprobante": fecha or "",
                "Tipo de Anulación": tipo_anulacion,
                "Estatus": status,
            })

        data_start = 12
        for i, data in enumerate(rows_data):
            row = data_start + i
            for name, value in data.items():
                col = header_map.get(name)
                if col is not None:
                    tf.write_cell(row, col, value)

        self._clear_remaining(tf, data_start, len(rows_data))
        return tf.save()

    def _export_608_csv(self, invoices, report_rnc=None, period=None) -> bytes:
        """608 en CSV (fallback rápido)."""
        output = io.StringIO()
        writer = csv.writer(output)
        rnc_value = self._only_digits(report_rnc) if report_rnc else ""
        period_value = period or self._derive_period(invoices) or ""
        writer.writerow(["FORMULARIO 608 - ANULACIONES DE COMPROBANTES"])
        writer.writerow([f"RNC Declarante: {rnc_value}", f"Período: {period_value}", f"Registros: {len(invoices)}"])
        writer.writerow([])
        writer.writerow(["NCF Anulado", "Tipo NCF", "Fecha Comprobante", "Fecha Anulación", "Proveedor", "RNC Proveedor", "Monto Original", "Estatus 608"])
        for inv in invoices:
            raw = self._parse_raw_data(inv.raw_extracted_data)
            ncf = inv.invoice_number or raw.get("invoice_number") or ""
            tipo_ncf = self._ncf_tipo(ncf)
            fecha = self._format_date(inv.invoice_date or raw.get("invoice_date"))
            fecha_anulacion = ""
            if inv.cancelled_at:
                fecha_anulacion = inv.cancelled_at.strftime('%Y%m%d')
            elif inv.deleted_at:
                fecha_anulacion = inv.deleted_at.strftime('%Y%m%d')
            elif inv.updated_at:
                fecha_anulacion = inv.updated_at.strftime('%Y%m%d')
            issues = []
            if not ncf:
                issues.append("Falta NCF")
            if not fecha_anulacion:
                issues.append("Falta fecha anulación")
            status = "; ".join(issues) if issues else "OK"
            writer.writerow([ncf, tipo_ncf, fecha or "", fecha_anulacion, inv.vendor_name or "",
                             self._only_digits(inv.vendor_tax_id) or "",
                             self._fmt_amount(self._to_number(inv.total_amount)) or '', status])
        return output.getvalue().encode("utf-8-sig")

    # ── DGII Official TXT generators ─────────────────────────────────────
    # Pipe-delimited .txt files matching the exact format the VBA macros
    # in the official DGII templates produce. Ready for direct upload
    # to the DGII Oficina Virtual (dgii.gov.do).
    #
    # Format spec extracted from:
    #   - Formato-de-Envio-606-(NG-07-2018-y-05-2019).xls  → modServicios.GenerarArchivo
    #   - Herramienta de Envio Formato 607.xls              → Formato607.GenerarArchivo
    #   - Herramienta de envio Formato 608.xls              → Formato608.cmdGenerarArchivo_Click

    def export_dgii_606_txt(self, invoices: List[Invoice], report_rnc: Optional[str] = None, period: Optional[str] = None) -> bytes:
        """Genera el archivo .txt oficial DGII 606 (Compras) — formato pipe-delimited.

        Estructura:
            Línea 1 (header): 606|RNC|PERIODO|CANT_REGISTROS
            Líneas 2+: RNC|TipoId|TipoBienes|NCF|NCFMod|FechaComprobante|FechaPago|
                        ServiciosFact|BienesFact|TotalFact|ITBISFact|ITBISRet|
                        ProporcionalidadITBIS|ITBISCosto|ITBISAdelantar|ITBISPercibido|
                        TipoRetISR|MontoRetRenta|ISRPercibido|ISC|OtrosImp|
                        PropinaLegal|FormaPago
        """
        rnc = self._only_digits(report_rnc) or ""
        period_value = period or self._derive_period(invoices) or ""
        lines = [f"606|{rnc}|{period_value}|{len(invoices)}"]

        for inv in invoices:
            fields = self._build_606_txt_fields(inv, report_rnc)
            lines.append("|".join(fields))

        # DGII macro: last line WITHOUT trailing newline (Print #1, strDetalle;)
        return "\r\n".join(lines).encode("utf-8")

    def export_dgii_607_txt(self, invoices: List[Invoice], report_rnc: Optional[str] = None, period: Optional[str] = None) -> bytes:
        """Genera el archivo .txt oficial DGII 607 (Ventas) — formato pipe-delimited.

        Estructura:
            Línea 1 (header): 607|RNC|PERIODO|CANT_REGISTROS
            Líneas 2+: RNC_Comprador|TipoId|NCF|NCFMod|TipoIngreso|FechaComp|FechaRet|
                        MontoFact|ITBISFact|ITBISRetTerceros|ITBISPercibido|
                        RetRentaTerceros|ISRPercibido|ISC|OtrosImp|PropinaLegal|
                        Efectivo|Cheque|Tarjeta|Credito|Bonos|Permuta|OtrasFormas
        """
        rnc = self._only_digits(report_rnc) or ""
        period_value = period or self._derive_period(invoices) or ""
        lines = [f"607|{rnc}|{period_value}|{len(invoices)}"]

        for inv in invoices:
            fields = self._build_607_txt_fields(inv)
            lines.append("|".join(fields))

        return "\r\n".join(lines).encode("utf-8")

    def export_dgii_608_txt(self, invoices: List[Invoice], report_rnc: Optional[str] = None, period: Optional[str] = None) -> bytes:
        """Genera el archivo .txt oficial DGII 608 (Anulaciones) — formato pipe-delimited.

        Estructura:
            Línea 1 (header): 608|RNC|PERIODO|CANT_REGISTROS
            Líneas 2+: NCF|FechaComprobante|TipoAnulacion
        """
        rnc = self._only_digits(report_rnc) or ""
        period_value = period or self._derive_period(invoices) or ""
        lines = [f"608|{rnc}|{period_value}|{len(invoices)}"]

        for inv in invoices:
            fields = self._build_608_txt_fields(inv)
            lines.append("|".join(fields))

        return "\r\n".join(lines).encode("utf-8")

    def _build_606_txt_fields(self, inv: Invoice, report_rnc: Optional[str] = None) -> List[str]:
        raw = self._parse_raw_data(inv.raw_extracted_data)
        goods_type = self._normalize_goods_type(inv.goods_services_type or raw.get("goods_services_type")) or ""
        ncf = inv.invoice_number or raw.get("invoice_number") or ""
        vendor_tax_id = inv.vendor_tax_id or raw.get("vendor_tax_id")
        if self._ncf_document_type(ncf) in {"13", "17", "43", "47"}:
            vendor_tax_id = report_rnc
        inv_rnc = self._only_digits(vendor_tax_id) or ""
        tipo_id = self._tipo_id_from_tax_id(inv_rnc) or ""
        ncf_modified = raw.get("ncf_modified") or ""
        fecha_comprobante = self._format_date(inv.invoice_date or raw.get("invoice_date")) or ""
        fecha_pago = self._format_date(raw.get("payment_date")) or ""

        total = self._to_number(inv.total_amount) or self._to_number(raw.get("total_amount"))
        tax = self._to_number(inv.tax_amount) or self._to_number(raw.get("tax_amount"))
        base = self._compute_base(total, tax, raw)
        amount_services, amount_goods = self._split_base(base, inv, goods_type, raw)
        total_facturado = (amount_services or 0.0) + (amount_goods or 0.0)
        itbis_facturado = self._to_number(tax) or 0.0
        itbis_retenido = self._to_number(raw.get("itbis_retenido"))
        itbis_proporcionalidad = self._to_number(raw.get("itbis_proporcionalidad"))
        itbis_llevado_costo = self._to_number(raw.get("itbis_llevado_costo"))
        itbis_adelantar = itbis_facturado - (itbis_llevado_costo or 0.0)
        if itbis_adelantar < 0:
            itbis_adelantar = 0.0
        itbis_percibido = self._to_number(raw.get("itbis_percibido"))
        isr_retention_type = self._normalize_isr_retention(raw.get("isr_retention_type"))
        isr_retention_amount = self._to_number(raw.get("isr_retention_amount"))
        isr_percibido = self._to_number(raw.get("isr_percibido"))
        isc_amount = self._to_number(raw.get("isc_amount"))
        other_taxes = self._to_number(raw.get("other_taxes"))
        legal_tip = self._to_number(raw.get("legal_tip"))
        payment_method = self._normalize_payment_method(raw.get("payment_method"))

        goods_type_code = goods_type[:2] if goods_type else ""
        isr_type_code = (isr_retention_type or "")[:2]
        payment_code = (payment_method or "")[:2]

        return [
            inv_rnc,
            tipo_id,
            goods_type_code,
            ncf,
            ncf_modified,
            fecha_comprobante,
            fecha_pago,
            self._txt_amount(amount_services, allow_zero=True),
            self._txt_amount(amount_goods, allow_zero=True),
            self._txt_amount(total_facturado, allow_zero=True),
            self._txt_amount(itbis_facturado, allow_zero=True),
            self._txt_amount(itbis_retenido),
            self._txt_amount(itbis_proporcionalidad),
            self._txt_amount(itbis_llevado_costo),
            self._txt_amount(itbis_adelantar, allow_zero=True),
            self._txt_amount(itbis_percibido),
            isr_type_code,
            self._txt_amount(isr_retention_amount),
            self._txt_amount(isr_percibido),
            self._txt_amount(isc_amount),
            self._txt_amount(other_taxes),
            self._txt_amount(legal_tip),
            payment_code,
        ]

    def _build_607_txt_fields(self, inv: Invoice) -> List[str]:
        raw = self._parse_raw_data(inv.raw_extracted_data)
        buyer_tax_id = inv.vendor_tax_id or raw.get("vendor_tax_id") or raw.get("buyer_tax_id")
        inv_rnc = self._only_digits(buyer_tax_id) or ""
        tipo_id = self._tipo_id_from_tax_id(inv_rnc) or ""
        ncf = inv.invoice_number or raw.get("invoice_number") or ""
        ncf_modified = raw.get("ncf_modified") or ""
        tipo_ingreso_raw = raw.get("tipo_ingreso") or "01"
        tipo_ingreso = tipo_ingreso_raw[-1] if len(tipo_ingreso_raw) >= 2 else tipo_ingreso_raw
        fecha = self._format_date(inv.invoice_date or raw.get("invoice_date")) or ""
        fecha_retencion = self._format_date(raw.get("payment_date")) or ""
        total = self._to_number(inv.total_amount) or self._to_number(raw.get("total_amount")) or 0.0
        tax = self._to_number(inv.tax_amount) or self._to_number(raw.get("tax_amount")) or 0.0
        amount_facturado = self._compute_base(total, tax, raw)
        itbis_retenido = self._to_number(raw.get("itbis_retenido")) or 0.0
        itbis_percibido = self._to_number(raw.get("itbis_percibido")) or 0.0
        ret_renta_terceros = self._txt_amount(self._to_number(raw.get("retencion_renta_terceros")))
        isr_percibido = self._txt_amount(self._to_number(raw.get("isr_percibido")))
        isc = self._txt_amount(self._to_number(raw.get("isc_amount")))
        otros_imp = self._txt_amount(self._to_number(raw.get("other_taxes")))
        propina = self._txt_amount(self._to_number(raw.get("legal_tip")))

        efectivo, cheque, tarjeta, credito, bonos, permuta, otras = self._payment_breakdown_607(
            raw.get("payment_method"),
            total,
        )

        return [
            inv_rnc,
            tipo_id,
            ncf,
            ncf_modified,
            tipo_ingreso,
            fecha,
            fecha_retencion,
            self._txt_amount(amount_facturado, allow_zero=True),
            self._txt_amount(tax, allow_zero=True),
            self._txt_amount(itbis_retenido),
            self._txt_amount(itbis_percibido),
            ret_renta_terceros,
            isr_percibido,
            isc,
            otros_imp,
            propina,
            self._txt_amount(efectivo),
            self._txt_amount(cheque),
            self._txt_amount(tarjeta),
            self._txt_amount(credito),
            self._txt_amount(bonos),
            self._txt_amount(permuta),
            self._txt_amount(otras),
        ]

    def _build_608_txt_fields(self, inv: Invoice) -> List[str]:
        raw = self._parse_raw_data(inv.raw_extracted_data)
        ncf = inv.invoice_number or raw.get("invoice_number") or ""
        fecha = self._format_date(inv.invoice_date or raw.get("invoice_date")) or ""
        tipo_anulacion = inv.cancellation_type or raw.get("cancellation_type") or "01"
        tipo_anulacion = tipo_anulacion[:2]
        return [ncf.strip(), fecha.strip(), tipo_anulacion.strip()]

    def _txt_amount(self, value, allow_zero: bool = False) -> str:
        """Format a numeric value for DGII .txt output.

        Official templates write required zero-valued cells as ``0``. Optional
        monetary fields remain blank when absent/zero.
        """
        if value is None:
            return ""
        val = round(float(value), 2)
        if val == 0 and not allow_zero:
            return ""
        # Remove trailing zeros: 1500.50 → "1500.5", 1500.00 → "1500"
        s = f"{val:.2f}"
        if "." in s:
            s = s.rstrip("0").rstrip(".")
        return s

    # ── Private helpers ─────────────────────────────────────────────────────

    def _parse_raw_data(self, raw):
        try:
            if raw:
                data = json.loads(raw)
                # Normalize retention key naming: total_* variant → short variant
                if "total_itbis_retenido" in data and "itbis_retenido" not in data:
                    data["itbis_retenido"] = data["total_itbis_retenido"]
                if "total_isr_retencion" in data and "isr_retention_amount" not in data:
                    data["isr_retention_amount"] = data["total_isr_retencion"]
                if "total_itbis_percepcion" in data and "itbis_percibido" not in data:
                    data["itbis_percibido"] = data["total_itbis_percepcion"]
                if "total_isr_percepcion" in data and "isr_percibido" not in data:
                    data["isr_percibido"] = data["total_isr_percepcion"]
                return data
        except Exception:
            pass
        return {}

    def _only_digits(self, value):
        if not value:
            return None
        return "".join([c for c in str(value) if c.isdigit()])

    def _tipo_id_from_tax_id(self, tax_id):
        if not tax_id:
            return None
        if len(tax_id) == 9:
            return "1"
        if len(tax_id) == 11:
            return "2"
        return None

    def _normalize_goods_type(self, value):
        if not value:
            return None
        digits = "".join([c for c in str(value) if c.isdigit()])
        if not digits:
            return None
        if len(digits) == 1:
            digits = f"0{digits}"
        valid = {f"{i:02d}" for i in range(1, 12)}
        return digits if digits in valid else None

    def _format_date(self, value):
        if not value:
            return None
        # datetime is a subclass of date — check it first
        if isinstance(value, datetime):
            return value.strftime('%Y%m%d')
        from datetime import date as _date
        if isinstance(value, _date):
            return value.strftime('%Y%m%d')
        try:
            if isinstance(value, str):
                value = value.strip()
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
                    try:
                        return datetime.strptime(value, fmt).strftime('%Y%m%d')
                    except ValueError:
                        continue
        except Exception:
            return None
        return None

    def _derive_period(self, invoices: List[Invoice]) -> Optional[str]:
        dates = [inv.invoice_date for inv in invoices if inv.invoice_date]
        if not dates:
            return None
        most_recent = max(dates)
        return most_recent.strftime('%Y%m')

    def _compute_base(self, total, tax, raw) -> float:
        base = None
        if total is not None and tax is not None:
            base = total - tax
        if base is None or base < 0:
            base = self._sum_line_items(raw.get("line_items"))
        return base if base is not None and base >= 0 else 0.0

    def _split_base(self, base, inv, goods_type, raw):
        amount_services = self._to_number(raw.get("services_amount"))
        amount_goods = self._to_number(raw.get("goods_amount"))
        if amount_services is None and amount_goods is None:
            amount_goods, amount_services = self._split_base_by_type(base, inv, goods_type)
        elif amount_services is None and amount_goods is not None:
            amount_services = max(base - amount_goods, 0.0)
        elif amount_goods is None and amount_services is not None:
            amount_goods = max(base - amount_services, 0.0)
        return amount_services or 0.0, amount_goods or 0.0

    def _to_number(self, value):
        try:
            if value is None or value == "":
                return None
            return float(value)
        except (TypeError, ValueError):
            return None

    def _fmt_amount(self, value, allow_zero=False):
        """Return a numeric value (float) for template cells, or None for blank."""
        if value is None:
            return None
        if value == 0 and not allow_zero:
            return None
        return round(float(value), 2)

    def _clear_remaining(self, tf: TemplateFiller, data_start: int, num_rows: int):
        """Clear example rows after our data."""
        tf.clear_from(data_start + num_rows)

    def _sum_line_items(self, items):
        if not isinstance(items, list):
            return None
        total = 0.0
        has_any = False
        for item in items:
            if not isinstance(item, dict):
                continue
            subtotal = self._to_number(item.get("subtotal"))
            if subtotal is None:
                continue
            total += subtotal
            has_any = True
        return total if has_any else None

    def _split_base_by_type(self, base, inv, goods_type):
        category = (inv.category or '').lower()
        goods_keywords = ['oficina', 'inventario', 'mercancia', 'mercancía', 'compras', 'equipos', 'activos', 'maquinaria']
        goods_types = {"04", "09", "10"}
        is_goods = goods_type in goods_types or any(k in category for k in goods_keywords)
        if is_goods:
            return base, 0.0
        return 0.0, base

    def _payment_breakdown_607(self, payment_method, total: float):
        """Return 607 sales buckets for columns 17-23.

        607 distributes the invoice total (base + taxes) across collection
        columns, while column 8 remains the tax-exclusive amount.
        """
        pm = self._normalize_payment_method(payment_method)
        payment_total = total or 0.0
        return (
            payment_total if pm == "01" else 0.0,
            payment_total if pm == "02" else 0.0,
            payment_total if pm == "03" else 0.0,
            payment_total if pm == "04" else 0.0,
            payment_total if pm == "06" else 0.0,
            payment_total if pm == "05" else 0.0,
            payment_total if pm == "07" else 0.0,
        )

    def _normalize_isr_retention(self, value):
        if value is None:
            return None
        raw = str(value).strip()
        if raw.isdigit():
            code = int(raw)
            return f"{code:02d}" if 1 <= code <= 9 else None
        text = raw.lower()
        mapping = {
            "alquiler": "01",
            "honorario": "02",
            "servicio": "02",
            "otras rentas": "03",
            "rentas presuntas": "04",
            "intereses pagados a personas juridicas": "05",
            "intereses pagados a personas jurídicas": "05",
            "intereses pagados a personas fisicas": "06",
            "intereses pagados a personas físicas": "06",
            "proveedores del estado": "07",
            "juegos telefonicos": "08",
            "juegos telefónicos": "08",
            "ganaderia": "09",
            "ganadería": "09"
        }
        for key, code in mapping.items():
            if key in text:
                return code
        return None

    def _normalize_payment_method(self, value):
        if value is None:
            return None
        raw = str(value).strip()
        if raw.isdigit():
            code = int(raw)
            return f"{code:02d}" if 1 <= code <= 10 else None
        text = raw.lower()
        if "efectivo" in text:
            return "01"
        if "cheque" in text or "transfer" in text or "depósito" in text or "deposito" in text:
            return "02"
        if "tarjeta" in text:
            return "03"
        if "crédito" in text or "credito" in text:
            return "04"
        if "bono" in text or "certificado" in text:
            return "06"
        if "permuta" in text:
            return "05"
        if "nota de crédito" in text or "nota de credito" in text:
            return "06"
        if "mixto" in text or "otra" in text:
            return "07"
        return None

    def _ncf_tipo(self, ncf: str) -> str:
        """Deriva el tipo de comprobante del prefijo del NCF."""
        if not ncf:
            return ""
        ncf = ncf.strip().upper()
        # e-CF: E31, E32, ... | B01, B02, etc.
        if ncf.startswith("E"):
            code = ncf[1:3] if len(ncf) >= 3 else ""
        elif ncf.startswith("B") or ncf.startswith("A"):
            code = ncf[1:3] if len(ncf) >= 3 else ""
        else:
            return ""
        tipo_map = {
            "01": "Factura de Crédito Fiscal",
            "02": "Factura de Consumo",
            "03": "Nota de Débito",
            "04": "Nota de Crédito",
            "11": "Proveedores Informales",
            "12": "Registro Único de Ingresos",
            "13": "Gastos Menores",
            "14": "Regímenes Especiales",
            "15": "Gubernamentales",
            "16": "Exportaciones",
            "31": "e-CF Crédito Fiscal",
            "32": "e-CF Consumo",
            "33": "e-CF Nota de Débito",
            "34": "e-CF Nota de Crédito",
            "41": "e-CF Compras",
            "43": "e-CF Gastos Menores",
            "44": "e-CF Regímenes Especiales",
            "45": "e-CF Gubernamentales",
            "46": "e-CF Exportaciones",
            "47": "e-CF Pagos al Exterior",
        }
        return tipo_map.get(code, f"Tipo {code}")

    def _ncf_document_type(self, ncf: str) -> str:
        if not ncf:
            return ""
        value = ncf.strip().upper()
        if len(value) == 13 and value.startswith("E"):
            return value[1:3]
        if len(value) == 11 and value[0] in {"B", "A", "P", "Q"}:
            return value[1:3]
        if len(value) == 19 and value[0] in {"A", "P", "Q"}:
            return value[9:11]
        return ""

    def _build_606_status(
        self,
        rnc,
        ncf,
        fecha_comprobante,
        fecha_pago,
        total_facturado,
        itbis_facturado,
        itbis_retenido,
        isr_retention_type,
        isr_retention_amount
    ):
        issues = []
        if not rnc:
            issues.append("Falta RNC/Cédula")
        if not ncf:
            issues.append("Falta NCF")
        if not fecha_comprobante:
            issues.append("Falta fecha comprobante")
        if (itbis_retenido or isr_retention_type or isr_retention_amount) and not fecha_pago:
            issues.append("Falta fecha pago")
        if total_facturado is None or total_facturado == 0:
            issues.append("Montos en cero")
        if itbis_facturado is not None and total_facturado is not None and itbis_facturado > 0 and total_facturado < 0:
            issues.append("Total inválido")
        return "; ".join(issues) if issues else "OK"
    def export_csv_generic(self, invoices: List[Invoice]) -> str:
        """CSV con datos fiscales relevantes para DGII / contabilidad."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(FISCAL_HEADERS)
        for inv in invoices:
            writer.writerow(_fiscal_row(inv))
        return output.getvalue()

    def export_quickbooks(self, invoices: List[Invoice]) -> str:
        """
        Formato compatible con importación de 'Bills' en QuickBooks Online.
        Headers: Bill No,Vendor,Transaction Date,Due Date,Total,Account,Line Amount,Line Description
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        headers = [
            'Bill No', 'Vendor', 'Transaction Date', 'Due Date', 
            'Total', 'Account', 'Line Amount', 'Line Description'
        ]
        writer.writerow(headers)
        
        for inv in invoices:
            # QuickBooks requiere mapeo de cuentas. Usamos la categoría o una cuenta por defecto
            account = inv.category or "Uncategorized Expense"
            
            # Asumimos fecha de vencimiento = fecha factura + 30 días si no hay dato
            date_str = inv.invoice_date.strftime('%m/%d/%Y') if inv.invoice_date else datetime.now().strftime('%m/%d/%Y')
            
            writer.writerow([
                inv.invoice_number or f"INV-{inv.id}",
                inv.vendor_name or "Unknown Vendor",
                date_str,
                date_str, # Due Date placeholder
                f"{inv.total_amount or 0:.2f}",
                account,
                f"{inv.total_amount or 0:.2f}", # Line Amount (simplificado a 1 linea)
                inv.description or "Services provided"
            ])
            
        return output.getvalue()

    def export_contaplus(self, invoices: List[Invoice]) -> str:
        """
        Formato simplificado tipo Diario para Sage/Contaplus (Asientos).
        Col: Fecha, Cuenta, Concepto, Debe, Haber, Documento
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        headers = ['Fecha', 'Cuenta', 'Concepto', 'Debe', 'Haber', 'Documento']
        writer.writerow(headers)
        
        for inv in invoices:
            date_str = inv.invoice_date.strftime('%d/%m/%Y') if inv.invoice_date else ''
            doc_ref = inv.invoice_number or f"DOC-{inv.id}"
            total = inv.total_amount or 0
            tax = inv.tax_amount or 0
            base = total - tax
            
            # Linea 1: Gasto (Base) -> Debe
            writer.writerow([
                date_str, 
                "60000000", # Cuenta genérica de compras (debería venir de settings)
                f"Fra. {inv.vendor_name}",
                f"{base:.2f}",
                "0.00",
                doc_ref
            ])
            
            # Linea 2: IVA -> Debe
            if tax > 0:
                writer.writerow([
                    date_str,
                    "47200000", # HP IVA Soportado
                    "IVA Soportado",
                    f"{tax:.2f}",
                    "0.00",
                    doc_ref
                ])
                
            # Linea 3: Proveedor -> Haber
            writer.writerow([
                date_str,
                "40000000", # Proveedores
                f"Fra. {inv.vendor_name}",
                "0.00",
                f"{total:.2f}",
                doc_ref
            ])
            
        return output.getvalue()

    def export_quickbooks_bills(self, invoices: List[Invoice]) -> str:
        """
        Formato Bills (QuickBooks Online) con columnas estándar.
        Headers: Bill No, Vendor, Transaction Date, Due Date, Account, Line Amount, Line Description, Total, Tax Code
        """
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            'Bill No', 'Vendor', 'Transaction Date', 'Due Date',
            'Account', 'Line Amount', 'Line Description', 'Total', 'Tax Code'
        ]
        writer.writerow(headers)

        for inv in invoices:
            date_val = inv.invoice_date or datetime.utcnow()
            date_str = date_val.strftime('%m/%d/%Y')
            due_str = (date_val + timedelta(days=30)).strftime('%m/%d/%Y')
            total = inv.total_amount or 0
            tax = inv.tax_amount or 0
            base = total - tax
            account = inv.category or "Expenses"

            writer.writerow([
                inv.invoice_number or f"INV-{inv.id}",
                inv.vendor_name or "Unknown Vendor",
                date_str,
                due_str,
                account,
                f"{base:.2f}",
                inv.description or "Services / Goods",
                f"{total:.2f}",
                ""
            ])

        return output.getvalue()

    def export_xero_bills(self, invoices: List[Invoice]) -> str:
        """
        Formato Bills (Xero) - CSV compatible con importación estándar.
        Headers: Contact Name, Invoice Number, Invoice Date, Due Date, Description, Quantity, Unit Amount, Account Code, Tax Type, Currency
        """
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            'Contact Name', 'Invoice Number', 'Invoice Date', 'Due Date',
            'Description', 'Quantity', 'Unit Amount', 'Account Code', 'Tax Type', 'Currency'
        ]
        writer.writerow(headers)

        for inv in invoices:
            date_val = inv.invoice_date or datetime.utcnow()
            date_str = date_val.strftime('%Y-%m-%d')
            due_str = (date_val + timedelta(days=30)).strftime('%Y-%m-%d')
            total = inv.total_amount or 0
            tax = inv.tax_amount or 0
            base = total - tax
            account = inv.category or "Expenses"

            writer.writerow([
                inv.vendor_name or "Unknown Vendor",
                inv.invoice_number or f"INV-{inv.id}",
                date_str,
                due_str,
                inv.description or "Services / Goods",
                "1",
                f"{base:.2f}",
                account,
                "",
                inv.currency or "DOP"
            ])

        return output.getvalue()

    def export_odoo_vendor_bills(self, invoices: List[Invoice]) -> str:
        """
        Formato CSV genérico para importación de facturas de proveedor en Odoo.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            'move_type', 'partner_id/name', 'invoice_date', 'invoice_date_due',
            'ref', 'currency_id/name',
            'invoice_line_ids/name', 'invoice_line_ids/quantity',
            'invoice_line_ids/price_unit', 'invoice_line_ids/account_id/name'
        ]
        writer.writerow(headers)

        for inv in invoices:
            date_val = inv.invoice_date or datetime.utcnow()
            date_str = date_val.strftime('%Y-%m-%d')
            due_str = (date_val + timedelta(days=30)).strftime('%Y-%m-%d')
            total = inv.total_amount or 0
            tax = inv.tax_amount or 0
            base = total - tax
            account = inv.category or "Expenses"

            writer.writerow([
                "in_invoice",
                inv.vendor_name or "Unknown Vendor",
                date_str,
                due_str,
                inv.invoice_number or f"INV-{inv.id}",
                inv.currency or "DOP",
                inv.description or "Services / Goods",
                "1",
                f"{base:.2f}",
                account
            ])

        return output.getvalue()

    def export_excel_generic(self, invoices: List[Invoice]) -> bytes:
        """Excel formateado con encabezados estilizados, columnas ajustadas y datos fiscales."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="B0B0B0"),
            right=Side(style="thin", color="B0B0B0"),
            top=Side(style="thin", color="B0B0B0"),
            bottom=Side(style="thin", color="B0B0B0"),
        )
        data_font = Font(name="Calibri", size=10)
        data_align = Alignment(vertical="center")
        number_align = Alignment(horizontal="right", vertical="center")
        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        headers = FISCAL_HEADERS
        col_widths = [14, 28, 20, 13, 10, 16, 14, 30, 16, 12, 16, 10, 14]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30

        for row_idx, inv in enumerate(invoices, 2):
            row_data = _fiscal_row(inv)
            is_alt = (row_idx % 2 == 0)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                if is_alt:
                    cell.fill = alt_fill
                if col_idx in (9, 10, 11):
                    cell.alignment = number_align
                else:
                    cell.alignment = data_align

        ws.auto_filter.ref = ws.dimensions

        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.tabColor = "1F4E79"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_txt(self, invoices: List[Invoice]) -> str:
        """TXT legible con datos fiscales relevantes."""
        lines: list[str] = []
        sep = "=" * 55
        for i, inv in enumerate(invoices, 1):
            base = (inv.total_amount or 0) - (inv.tax_amount or 0)
            lines.append(sep)
            lines.append(f"FACTURA #{i}")
            lines.append(sep)
            lines.append(f"RNC:\t\t{inv.vendor_tax_id or '—'}")
            lines.append(f"Proveedor:\t{inv.vendor_name or '—'}")
            lines.append(f"NCF:\t\t{inv.invoice_number or '—'}")
            lines.append(f"Fecha:\t\t{inv.invoice_date.strftime('%d/%m/%Y') if inv.invoice_date else '—'}")
            tipo = {"income": "Ingreso", "expense": "Gasto"}.get(inv.transaction_type or "", "")
            lines.append(f"Tipo:\t\t{tipo}" if tipo else "")
            lines.append(f"Categoría:\t{inv.category or '—'}")
            lines.append(f"Tipo B/S:\t{inv.goods_services_type or '—'}")
            lines.append(f"Descripción:\t{inv.description or '—'}")
            lines.append(f"Base:\t\t{base:,.2f}")
            lines.append(f"ITBIS:\t\t{inv.tax_amount or 0:,.2f}")
            lines.append(f"Total:\t\t{inv.total_amount or 0:,.2f}")
            lines.append(f"Moneda:\t\t{inv.currency or 'DOP'}")
            lines.append("")
        lines.append(sep)
        lines.append(f"Total facturas: {len(invoices)}")
        total = sum(inv.total_amount or 0 for inv in invoices)
        lines.append(f"Suma total:\t{total:,.2f}")
        lines.append(sep)
        return "\n".join(lines)

    def export_json(self, invoices: List[Invoice]) -> str:
        """JSON con datos fiscales relevantes — sin fugas del esquema interno."""
        data = []
        for inv in invoices:
            base = (inv.total_amount or 0) - (inv.tax_amount or 0)
            data.append({
                "rnc": inv.vendor_tax_id or "",
                "proveedor": inv.vendor_name or "",
                "ncf": inv.invoice_number or "",
                "fecha": inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
                "tipo": {"income": "Ingreso", "expense": "Gasto"}.get(inv.transaction_type or "", inv.transaction_type or ""),
                "categoria": inv.category or "",
                "tipo_bien_serv": inv.goods_services_type or "",
                "descripcion": inv.description or "",
                "base_imponible": round(base, 2),
                "itbis": round(inv.tax_amount or 0, 2),
                "total": round(inv.total_amount or 0, 2),
                "moneda": inv.currency or "DOP",
                "estado": "Procesado" if inv.processed else "Pendiente",
            })
        return json.dumps(data, indent=2, ensure_ascii=False)
