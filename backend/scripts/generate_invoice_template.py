"""
Generate a modern, UX-focused Excel template for bulk invoice import.
Run: python scripts/generate_invoice_template.py
Output: templates/invoice_import_template.xlsx
"""

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, NamedStyle, Protection, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ─── Brand Colors (Fintral sky-blue palette) ────────────────────────────
SKY_950 = "082F49"   # Darkest
SKY_700 = "0369A1"
SKY_500 = "0EA5E9"   # Primary
SKY_400 = "38BDF8"
SKY_300 = "7DD3FC"
SKY_200 = "BAE6FD"
SKY_100 = "E0F2FE"
SKY_50  = "F0F9FF"

ZINC_950 = "09090B"
ZINC_700 = "3F3F46"
ZINC_500 = "71717A"
ZINC_300 = "D4D4D8"
ZINC_200 = "E4E4E7"
ZINC_100 = "F4F4F5"
ZINC_50  = "FAFAFA"
WHITE    = "FFFFFF"

GREEN_50  = "F0FDF4"
GREEN_600 = "16A34A"
AMBER_50  = "FFFBEB"
AMBER_600 = "D97706"
RED_50    = "FEF2F2"
RED_600   = "DC2626"

# ─── Reusable Styles ────────────────────────────────────────────────────
FONT_TITLE = Font(name="Aptos", size=18, bold=True, color=SKY_950)
FONT_SUBTITLE = Font(name="Aptos", size=11, color=ZINC_500)
FONT_SECTION = Font(name="Aptos", size=13, bold=True, color=SKY_700)
FONT_BODY = Font(name="Aptos", size=10, color=ZINC_700)
FONT_BODY_BOLD = Font(name="Aptos", size=10, bold=True, color=ZINC_950)
FONT_SMALL = Font(name="Aptos", size=9, color=ZINC_500)
FONT_HEADER = Font(name="Aptos", size=10, bold=True, color=WHITE)
FONT_REQUIRED_HEADER = Font(name="Aptos", size=10, bold=True, color=WHITE)
FONT_OPTIONAL_HEADER = Font(name="Aptos", size=10, bold=False, color=WHITE)
FONT_EXAMPLE = Font(name="Aptos", size=10, color=ZINC_500, italic=True)
FONT_TAG = Font(name="Aptos", size=9, bold=True, color=SKY_700)
FONT_LINK = Font(name="Aptos", size=10, color=SKY_500, underline="single")
FONT_CHECK = Font(name="Aptos", size=10, color=GREEN_600)
FONT_WARN = Font(name="Aptos", size=10, color=AMBER_600)

FILL_WHITE = PatternFill("solid", fgColor=WHITE)
FILL_SKY_50 = PatternFill("solid", fgColor=SKY_50)
FILL_SKY_100 = PatternFill("solid", fgColor=SKY_100)
FILL_SKY_500 = PatternFill("solid", fgColor=SKY_500)
FILL_SKY_950 = PatternFill("solid", fgColor=SKY_950)
FILL_ZINC_50 = PatternFill("solid", fgColor=ZINC_50)
FILL_ZINC_100 = PatternFill("solid", fgColor=ZINC_100)
FILL_GREEN_50 = PatternFill("solid", fgColor=GREEN_50)
FILL_AMBER_50 = PatternFill("solid", fgColor=AMBER_50)
FILL_RED_50 = PatternFill("solid", fgColor=RED_50)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin", color=ZINC_200),
    right=Side(style="thin", color=ZINC_200),
    top=Side(style="thin", color=ZINC_200),
    bottom=Side(style="thin", color=ZINC_200),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color=SKY_700),
    right=Side(style="thin", color=SKY_700),
    top=Side(style="thin", color=SKY_700),
    bottom=Side(style="medium", color=SKY_400),
)
NO_BORDER = Border()


def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Helper to set a cell with all formatting at once."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def fill_row(ws, row, start_col, end_col, fill, border=None):
    """Fill a range of cells with a background color."""
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        if border:
            cell.border = border


def build_bienvenida(wb):
    """Sheet 1: Bienvenida — clean onboarding page."""
    ws = wb.active
    ws.title = "Bienvenida"
    ws.sheet_properties.tabColor = SKY_500

    # Page setup
    ws.sheet_format.defaultRowHeight = 16
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 4   # margin
    ws.column_dimensions["B"].width = 3   # icon col
    ws.column_dimensions["C"].width = 50  # content
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 4   # margin

    # White background
    for r in range(1, 50):
        fill_row(ws, r, 1, 6, FILL_WHITE)

    # ── Logo area ──
    ws.row_dimensions[2].height = 36
    set_cell(ws, 2, 2, "▰▰▰", Font(name="Aptos", size=16, bold=True, color=SKY_500), FILL_WHITE, ALIGN_LEFT)
    set_cell(ws, 2, 3, "Fintral", FONT_TITLE, FILL_WHITE, Alignment(horizontal="left", vertical="center"))

    ws.row_dimensions[3].height = 20
    set_cell(ws, 3, 3, "Plantilla de Importación de Facturas", FONT_SUBTITLE, FILL_WHITE, ALIGN_LEFT)

    # ── Divider ──
    ws.row_dimensions[4].height = 6
    for c in range(2, 6):
        ws.cell(row=4, column=c).border = Border(bottom=Side(style="thin", color=SKY_200))

    # ── Quick start steps ──
    ws.row_dimensions[6].height = 24
    set_cell(ws, 6, 2, None, None, FILL_WHITE)
    set_cell(ws, 6, 3, "Cómo usar esta plantilla", FONT_SECTION, FILL_WHITE, ALIGN_LEFT)

    steps = [
        ("①", "Vaya a la hoja «Facturas»", "Allí ingresará toda la información."),
        ("②", "Complete las columnas obligatorias", "Marcadas con ★ en el encabezado."),
        ("③", "Use los desplegables", "Moneda, Tipo B/S y Forma de Pago tienen listas predefinidas."),
        ("④", "No deje filas vacías", "El sistema se detiene al encontrar una fila sin RNC."),
        ("⑤", "Guarde como .xlsx y suba al sistema", "Menú: Pipeline → Importar XLSX."),
    ]

    row = 8
    for num, title, desc in steps:
        ws.row_dimensions[row].height = 32
        set_cell(ws, row, 2, num, Font(name="Aptos", size=14, color=SKY_500), FILL_WHITE, ALIGN_CENTER)
        set_cell(ws, row, 3, title, FONT_BODY_BOLD, FILL_WHITE, ALIGN_LEFT)
        set_cell(ws, row, 4, desc, FONT_SMALL, FILL_WHITE, ALIGN_LEFT)
        row += 1

    # ── Field reference table ──
    row += 1
    ws.row_dimensions[row].height = 24
    set_cell(ws, row, 3, "Referencia de Campos", FONT_SECTION, FILL_WHITE, ALIGN_LEFT)
    row += 1

    # Table header
    ws.row_dimensions[row].height = 26
    ref_headers = [("Campo", 2), ("Obligatorio", 3), ("Formato", 4), ("Ejemplo", 5)]
    for label, col in ref_headers:
        set_cell(ws, row, col, label, FONT_HEADER, FILL_SKY_950, ALIGN_CENTER, HEADER_BORDER)
    row += 1

    fields = [
        ("RNC Proveedor",     "★ Sí",    "9 u 11 dígitos",         "501201234"),
        ("Razón Social",      "★ Sí",    "Texto libre",            "Servicios Técnicos SRL"),
        ("NCF",               "★ Sí",    "B/E + 8-10 dígitos",     "B0100001001"),
        ("Fecha Factura",     "★ Sí",    "AAAA-MM-DD",             "2026-01-15"),
        ("Fecha Pago",        "Opcional", "AAAA-MM-DD",             "2026-02-15"),
        ("Monto Total",       "★ Sí",    "Número > 0",             "5,000.00"),
        ("ITBIS",             "Opcional", "Número ≥ 0",             "750.00"),
        ("Tipo B/S (606)",    "Opcional", "01–11 (desplegable)",    "02"),
        ("Forma de Pago",     "Opcional", "1–7 (desplegable)",      "2"),
        ("Moneda",            "Opcional", "DOP / USD / EUR",        "DOP"),
        ("Tipo Transacción",  "Opcional", "income / expense",       "expense"),
        ("Categoría",         "Opcional", "Texto libre",            "servicios"),
        ("Descripción",       "Opcional", "Texto libre",            "Mantenimiento mensual"),
    ]

    for campo, req, fmt, ejemplo in fields:
        ws.row_dimensions[row].height = 22
        is_required = "★" in req
        bg = FILL_SKY_50 if is_required else FILL_WHITE
        font_req = Font(name="Aptos", size=9, bold=True, color=GREEN_600) if is_required else FONT_SMALL

        set_cell(ws, row, 2, campo, FONT_BODY_BOLD if is_required else FONT_BODY, bg, ALIGN_LEFT, THIN_BORDER)
        set_cell(ws, row, 3, req, font_req, bg, ALIGN_CENTER, THIN_BORDER)
        set_cell(ws, row, 4, fmt, FONT_SMALL, bg, ALIGN_LEFT, THIN_BORDER)
        set_cell(ws, row, 5, ejemplo, FONT_EXAMPLE, bg, ALIGN_LEFT, THIN_BORDER)
        row += 1

    # ── DGII codes reference ──
    row += 1
    ws.row_dimensions[row].height = 24
    set_cell(ws, row, 3, "Códigos DGII 606 — Tipo B/S", FONT_SECTION, FILL_WHITE, ALIGN_LEFT)
    row += 1

    dgii_codes = [
        ("01", "Gastos de personal"),
        ("02", "Gastos por trabajos, suministros y servicios"),
        ("03", "Arrendamientos"),
        ("04", "Gastos de activos fijos"),
        ("05", "Gastos de representación"),
        ("06", "Otras deducciones admitidas"),
        ("07", "Gastos financieros"),
        ("08", "Gastos extraordinarios"),
        ("09", "Compras y gastos (costo de venta)"),
        ("10", "Adquisiciones de activos"),
        ("11", "Gastos de seguros"),
    ]

    for code, desc in dgii_codes:
        ws.row_dimensions[row].height = 20
        set_cell(ws, row, 2, code, Font(name="Aptos Narrow", size=10, bold=True, color=SKY_500), FILL_WHITE, ALIGN_CENTER)
        set_cell(ws, row, 3, desc, FONT_BODY, FILL_WHITE, ALIGN_LEFT)
        row += 1

    # ── Payment methods ──
    row += 1
    ws.row_dimensions[row].height = 24
    set_cell(ws, row, 3, "Formas de Pago (DGII)", FONT_SECTION, FILL_WHITE, ALIGN_LEFT)
    row += 1

    payments = [
        ("1", "Efectivo"),
        ("2", "Cheque / Transferencia / Depósito"),
        ("3", "Tarjeta crédito / débito"),
        ("4", "Compra a crédito"),
        ("5", "Permuta"),
        ("6", "Notas de crédito"),
        ("7", "Mixto"),
    ]

    for code, desc in payments:
        ws.row_dimensions[row].height = 20
        set_cell(ws, row, 2, code, Font(name="Aptos Narrow", size=10, bold=True, color=SKY_500), FILL_WHITE, ALIGN_CENTER)
        set_cell(ws, row, 3, desc, FONT_BODY, FILL_WHITE, ALIGN_LEFT)
        row += 1

    # ── Footer ──
    row += 2
    set_cell(ws, row, 3, "Fintral — Financial Infrastructure", Font(name="Aptos", size=9, color=ZINC_300), FILL_WHITE, ALIGN_LEFT)

    # Protect sheet (read-only instructions)
    ws.protection.sheet = True
    ws.protection.password = ""  # visual protection only

    # Hide gridlines
    ws.sheet_view.showGridLines = False

    return ws


def build_facturas(wb):
    """Sheet 2: Facturas — the data entry sheet."""
    ws = wb.create_sheet("Facturas")
    ws.sheet_properties.tabColor = GREEN_600

    # ── Column config ──
    columns = [
        # (letter, width, header_label, is_required, number_format)
        ("A", 18,  "★ RNC Proveedor",    True,  "@"),
        ("B", 30,  "★ Razón Social",     True,  "@"),
        ("C", 20,  "★ NCF",              True,  "@"),
        ("D", 16,  "★ Fecha Factura",    True,  "YYYY-MM-DD"),
        ("E", 16,  "Fecha Pago",         False, "YYYY-MM-DD"),
        ("F", 16,  "★ Monto Total",      True,  "#,##0.00"),
        ("G", 14,  "ITBIS",              False, "#,##0.00"),
        ("H", 14,  "Tipo B/S (606)",     False, "@"),
        ("I", 16,  "Forma de Pago",      False, "@"),
        ("J", 10,  "Moneda",             False, "@"),
        ("K", 18,  "Tipo Transacción",   False, "@"),
        ("L", 18,  "Categoría",          False, "@"),
        ("M", 40,  "Descripción",        False, "@"),
    ]

    for letter, width, _, _, _ in columns:
        ws.column_dimensions[letter].width = width

    # ── Header row ──
    ws.row_dimensions[1].height = 36
    for idx, (_, _, label, is_req, _) in enumerate(columns, 1):
        font = FONT_REQUIRED_HEADER if is_req else FONT_OPTIONAL_HEADER
        fill = PatternFill("solid", fgColor=SKY_950) if is_req else PatternFill("solid", fgColor=SKY_700)
        set_cell(ws, 1, idx, label, font, fill, ALIGN_CENTER, HEADER_BORDER)

    # ── Sub-header with format hints ──
    ws.row_dimensions[2].height = 22
    hints = [
        "9 u 11 dígitos",  # A
        "Nombre oficial",   # B
        "B01/E31...",       # C
        "AAAA-MM-DD",       # D
        "AAAA-MM-DD",       # E
        "Ej: 5000.00",     # F
        "Ej: 750.00",      # G
        "01–11",            # H
        "1–7",              # I
        "DOP/USD",          # J
        "expense/income",   # K
        "Texto libre",      # L
        "Texto libre",      # M
    ]
    for idx, hint in enumerate(hints, 1):
        set_cell(ws, 2, idx, hint, FONT_SMALL, FILL_SKY_100, ALIGN_CENTER, THIN_BORDER)

    # ── Example row (row 3, light gray italic) ──
    ws.row_dimensions[3].height = 24
    example_data = [
        "501201234",
        "Servicios Técnicos SRL",
        "B0100001001",
        "2026-01-15",
        "2026-01-31",
        5000.00,
        750.00,
        "02",
        "2",
        "DOP",
        "expense",
        "servicios",
        "Mantenimiento sistema mensual",
    ]
    for idx, val in enumerate(example_data, 1):
        set_cell(ws, 3, idx, val, FONT_EXAMPLE, FILL_ZINC_50, ALIGN_LEFT, THIN_BORDER)

    # ── Data entry rows (4–1003) — 1000 rows pre-formatted ──
    DATA_START = 4
    DATA_END = 1003

    for row in range(DATA_START, DATA_END + 1):
        ws.row_dimensions[row].height = 22
        for idx, (_, _, _, is_req, num_fmt) in enumerate(columns, 1):
            bg = FILL_WHITE
            cell = set_cell(ws, row, idx, None, FONT_BODY, bg, ALIGN_LEFT, THIN_BORDER, num_fmt)

    # ── Required-column highlighting: soft sky-blue on required cols ──
    req_cols = [1, 2, 3, 4, 6]  # A, B, C, D, F
    for row in range(DATA_START, DATA_END + 1):
        for col in req_cols:
            ws.cell(row=row, column=col).fill = FILL_SKY_50

    # ── Zebra striping (subtle) on optional cols ──
    opt_cols = [5, 7, 8, 9, 10, 11, 12, 13]
    for row in range(DATA_START, DATA_END + 1):
        if row % 2 == 0:
            for col in opt_cols:
                ws.cell(row=row, column=col).fill = FILL_ZINC_50

    # ── Data Validations ──

    # Moneda (J)
    dv_moneda = DataValidation(
        type="list",
        formula1='"DOP,USD,EUR,MXN,COP,GBP,CAD"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Moneda inválida",
        error="Seleccione una moneda de la lista: DOP, USD, EUR, MXN, COP, GBP, CAD",
        showInputMessage=True,
        promptTitle="Moneda",
        prompt="Seleccione la moneda de la factura.\nPor defecto: DOP"
    )
    dv_moneda.sqref = f"J{DATA_START}:J{DATA_END}"
    ws.add_data_validation(dv_moneda)

    # Tipo B/S (H)
    dv_tipo = DataValidation(
        type="list",
        formula1='"01,02,03,04,05,06,07,08,09,10,11"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Tipo B/S inválido",
        error="Use un código del 01 al 11 según la tabla DGII 606.\nVea la hoja Bienvenida para referencia.",
        showInputMessage=True,
        promptTitle="Tipo Bienes/Servicios",
        prompt="01=Personal, 02=Servicios, 03=Arrend.,\n04=Activos fijos, 05=Representación,\n06=Otras, 07=Financieros, 08=Extraord.,\n09=Costo venta, 10=Adq. activos, 11=Seguros"
    )
    dv_tipo.sqref = f"H{DATA_START}:H{DATA_END}"
    ws.add_data_validation(dv_tipo)

    # Forma de Pago (I)
    dv_pago = DataValidation(
        type="list",
        formula1='"1,2,3,4,5,6,7"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Forma de pago inválida",
        error="Use un código del 1 al 7.\n1=Efectivo, 2=Cheque/Trans., 3=Tarjeta,\n4=Crédito, 5=Permuta, 6=NC, 7=Mixto",
        showInputMessage=True,
        promptTitle="Forma de Pago",
        prompt="1=Efectivo, 2=Cheque/Transferencia,\n3=Tarjeta, 4=Crédito, 5=Permuta,\n6=Nota de crédito, 7=Mixto"
    )
    dv_pago.sqref = f"I{DATA_START}:I{DATA_END}"
    ws.add_data_validation(dv_pago)

    # Tipo Transacción (K)
    dv_trans = DataValidation(
        type="list",
        formula1='"expense,income"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Tipo inválido",
        error="Use 'expense' para gastos o 'income' para ingresos.",
        showInputMessage=True,
        promptTitle="Tipo de Transacción",
        prompt="expense = Gasto (compra)\nincome = Ingreso (venta)"
    )
    dv_trans.sqref = f"K{DATA_START}:K{DATA_END}"
    ws.add_data_validation(dv_trans)

    # Monto Total > 0 (F)
    dv_monto = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Monto inválido",
        error="El monto total debe ser un número mayor que 0.",
        showInputMessage=True,
        promptTitle="Monto Total",
        prompt="Ingrese el monto total de la factura\nincluyendo impuestos."
    )
    dv_monto.sqref = f"F{DATA_START}:F{DATA_END}"
    ws.add_data_validation(dv_monto)

    # ITBIS >= 0 (G)
    dv_itbis = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="ITBIS inválido",
        error="El ITBIS debe ser un número mayor o igual a 0.",
        showInputMessage=True,
        promptTitle="ITBIS",
        prompt="Monto del ITBIS (impuesto).\nDeje vacío si no aplica."
    )
    dv_itbis.sqref = f"G{DATA_START}:G{DATA_END}"
    ws.add_data_validation(dv_itbis)

    # ── Conditional formatting: highlight empty required cells ──
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    red_font = Font(color=RED_600)

    # If a row has ANY data but a required col is empty → red
    for col_letter in ["A", "B", "C", "D", "F"]:
        rule = CellIsRule(
            operator="equal",
            formula=['""'],
            fill=red_fill,
            font=red_font,
        )
        ws.conditional_formatting.add(
            f"{col_letter}{DATA_START}:{col_letter}{DATA_END}",
            rule
        )

    # ── Freeze panes: keep header + hints visible ──
    ws.freeze_panes = "A4"

    # ── Auto-filter on header row ──
    ws.auto_filter.ref = f"A1:M{DATA_END}"

    # Hide gridlines for cleaner look
    ws.sheet_view.showGridLines = False

    # Print setup
    ws.print_title_rows = "1:2"

    return ws


def main():
    wb = openpyxl.Workbook()

    build_bienvenida(wb)
    build_facturas(wb)

    output_path = "templates/invoice_import_template.xlsx"
    wb.save(output_path)
    print(f"✅ Template saved to {output_path}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Data rows: 1000 (rows 4–1003)")
    print(f"   Validations: Moneda, Tipo B/S, Forma Pago, Tipo Transacción, Monto, ITBIS")


if __name__ == "__main__":
    main()
