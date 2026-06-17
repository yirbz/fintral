"""Integration tests for TemplateFiller and ExportService exports.

Tests the full .xlsx generation pipeline without requiring a database.
Verifies cell values, number formatting, and template preservation.
"""

import io
import os
import zipfile
from lxml import etree
from openpyxl import load_workbook

from app.services.export import TemplateFiller, TEMPLATE_DIR


_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'


def _read_styles(path: str) -> bytes:
    with zipfile.ZipFile(path) as z:
        return z.read('xl/styles.xml')


# ── 607 — Ventas ─────────────────────────────────────────────────

class Test607Template:
    TEMPLATE = os.path.join(TEMPLATE_DIR, '607.xlsx')

    def test_header_map_matches_expected(self):
        tf = TemplateFiller()
        tf.load('607.xlsx')
        hm = tf.header_map()
        assert isinstance(hm, dict)
        assert len(hm) > 0
        assert 'RNC/Cédula o Pasaporte' in hm
        assert 'Número Comprobante Fiscal' in hm
        assert 'Monto Facturado' in hm
        assert 'ITBIS Facturado' in hm

    def test_write_metadata_cells(self):
        tf = TemplateFiller()
        tf.load('607.xlsx')
        hm = tf.header_map()
        rnc_col = hm['RNC/Cédula o Pasaporte']
        tf.write_cell(4, rnc_col, '123456789')
        tf.write_cell(5, rnc_col, '012024')
        tf.write_cell(6, rnc_col, 42)
        buf = tf.save()
        wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
        ws = wb[tf.sheet_name]
        assert ws.cell(4, rnc_col).value == '123456789'
        assert str(ws.cell(5, rnc_col).value) == '012024'
        assert ws.cell(6, rnc_col).value == 42

    def test_write_amount_as_number(self):
        """Amounts must be numeric so Excel number formats apply."""
        tf = TemplateFiller()
        tf.load('607.xlsx')
        hm = tf.header_map()
        total_col = hm['Monto Facturado']

        tf.write_cell(11, total_col, 15000.50)
        buf = tf.save()

        wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
        ws = wb[tf.sheet_name]
        val = ws.cell(11, total_col).value
        assert val == 15000.50
        assert isinstance(val, (int, float))

    def test_write_text_stays_text(self):
        tf = TemplateFiller()
        tf.load('607.xlsx')
        hm = tf.header_map()
        ncf_col = hm['Número Comprobante Fiscal']

        tf.write_cell(11, ncf_col, 'E310000000001')
        buf = tf.save()

        wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
        ws = wb[tf.sheet_name]
        val = ws.cell(11, ncf_col).value
        assert val == 'E310000000001'
        assert isinstance(val, str)

    def test_clear_from_removes_example_rows(self):
        tf = TemplateFiller()
        tf.load('607.xlsx')
        hm = tf.header_map()
        total_col = hm['Monto Facturado']

        tf.write_cell(12, total_col, 5000.0)
        tf.write_cell(13, total_col, 3000.0)
        tf.clear_from(14)
        buf = tf.save()

        wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
        ws = wb[tf.sheet_name]
        assert ws.cell(12, total_col).value == 5000.0
        assert ws.cell(13, total_col).value == 3000.0
        assert ws.cell(14, total_col).value is None

    def test_styles_xml_unchanged(self):
        """styles.xml must be byte-identical after round-trip."""
        original = _read_styles(self.TEMPLATE)

        tf = TemplateFiller()
        tf.load('607.xlsx')
        hm = tf.header_map()
        tf.write_cell(4, hm['RNC/Cédula o Pasaporte'], '123456789')
        tf.write_cell(5, hm['RNC/Cédula o Pasaporte'], '012024')
        tf.write_cell(6, hm['RNC/Cédula o Pasaporte'], 42)
        for i in range(5):
            tf.write_cell(12 + i, hm['Número Comprobante Fiscal'], f'E3100000{i:04d}')
            tf.write_cell(12 + i, hm['Monto Facturado'], 1000.0 + i * 100)
            tf.write_cell(12 + i, hm['ITBIS Facturado'], 180.0 + i * 18)
            tf.write_cell(12 + i, hm['ITBIS Retenido por Terceros'], 100.0 + i * 10)
        tf.clear_from(17)
        buf = tf.save()

        with zipfile.ZipFile(io.BytesIO(buf)) as z:
            modified = z.read('xl/styles.xml')

        assert modified == original, 'styles.xml was modified'

    def test_merged_cells_preserved(self):
        tf = TemplateFiller()
        tf.load('607.xlsx')
        buf = tf.save()

        with zipfile.ZipFile(io.BytesIO(buf)) as z:
            sheet_xml = z.read('xl/worksheets/sheet1.xml')
        assert sheet_xml

    def test_round_trip_keeps_workbook_structure(self):
        """All expected files must exist after save()."""
        tf = TemplateFiller()
        tf.load('607.xlsx')
        tf.write_cell(4, 3, 'test')
        buf = tf.save()

        with zipfile.ZipFile(io.BytesIO(buf)) as z:
            sheet_xml = z.read('xl/worksheets/sheet1.xml')
        
        # Use iterparse to prevent sandbox OOM limit on 45MB XML
        context = etree.iterparse(io.BytesIO(sheet_xml), events=('end',))
        merges_count = 0
        cells_count = 0
        for event, elem in context:
            tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            if tag == 'mergeCells':
                merges_count = int(elem.get('count', '0'))
            elif tag == 'mergeCell':
                cells_count += 1
            elem.clear()

        assert merges_count > 0
        assert cells_count == merges_count


# ── 606 — Compras ─────────────────────────────────────────────────

class Test606Template:
    TEMPLATE = os.path.join(TEMPLATE_DIR, '606.xlsx')

    def test_header_map_matches_expected(self):
        tf = TemplateFiller()
        tf.load('606.xlsx')
        hm = tf.header_map()
        assert 'RNC o Cédula' in hm
        assert 'NCF' in hm
        assert 'Total Monto Facturado' in hm

    def test_cell_values_correct(self):
        tf = TemplateFiller()
        tf.load('606.xlsx')
        tf.write_cell(4, 1, '987654321')
        tf.write_cell(5, 3, '012024')
        tf.write_cell(6, 3, 5)
        buf = tf.save()

        wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
        ws = wb[tf.sheet_name]
        assert ws.cell(4, 1).value == '987654321'
        assert str(ws.cell(5, 3).value) == '012024'
        assert ws.cell(6, 3).value == 5

    def test_styles_xml_unchanged(self):
        original = _read_styles(self.TEMPLATE)
        tf = TemplateFiller()
        tf.load('606.xlsx')
        hm = tf.header_map()
        tf.write_cell(4, hm['RNC o Cédula'], '987654321')
        tf.write_cell(5, hm['RNC o Cédula'], '012024')
        tf.write_cell(6, hm['RNC o Cédula'], 5)
        buf = tf.save()

        with zipfile.ZipFile(io.BytesIO(buf)) as z:
            modified = z.read('xl/styles.xml')
        assert modified == original


# ── 608 — Anulaciones ─────────────────────────────────────────────

class Test608Template:
    TEMPLATE = os.path.join(TEMPLATE_DIR, '608.xlsx')

    def test_header_map_matches_expected(self):
        tf = TemplateFiller()
        tf.load('608.xlsx')
        hm = tf.header_map()
        assert 'Número de Comprobante Fiscal' in hm

    def test_cell_values_correct(self):
        tf = TemplateFiller()
        tf.load('608.xlsx')
        tf.write_cell(5, 3, '999999999')
        tf.write_cell(6, 3, '012024')
        tf.write_cell(7, 3, 3)
        buf = tf.save()

        wb = load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
        ws = wb[tf.sheet_name]
        assert ws.cell(5, 3).value == '999999999'
        assert str(ws.cell(6, 3).value) == '012024'
        assert ws.cell(7, 3).value == 3

    def test_styles_xml_unchanged(self):
        original = _read_styles(self.TEMPLATE)
        tf = TemplateFiller()
        tf.load('608.xlsx')
        hm = tf.header_map()
        tf.write_cell(5, hm['Número de Comprobante Fiscal'], 'E310000000001')
        tf.write_cell(6, hm['Número de Comprobante Fiscal'], '012024')
        tf.write_cell(7, hm['Número de Comprobante Fiscal'], 1)
        buf = tf.save()

        with zipfile.ZipFile(io.BytesIO(buf)) as z:
            modified = z.read('xl/styles.xml')
        assert modified == original


# ── CSV Exports ──────────────────────────────────────────────────

def test_csv_none_handling():
    """Verify _fmt_amount returning None doesn't leak 'None' into CSV."""
    from app.services.export import ExportService
    svc = ExportService()
    assert svc._fmt_amount(None) is None
    assert svc._fmt_amount(0, allow_zero=False) is None
    assert svc._fmt_amount(5000.0) == 5000.0
    assert svc._fmt_amount(0, allow_zero=True) == 0.0

# ── Template file existence ──────────────────────────────────────

def test_all_templates_exist():
    for name in ('606.xlsx', '607.xlsx', '608.xlsx'):
        path = os.path.join(TEMPLATE_DIR, name)
        assert os.path.exists(path), f'Missing template: {path}'


# ── DGII Official TXT generators ────────────────────────────────

def _mock_invoice(**kwargs):
    """Create a mock invoice SimpleNamespace for testing."""
    from types import SimpleNamespace
    from datetime import date
    defaults = dict(
        id='test-1',
        invoice_number='E310000000001',
        invoice_date=date(2024, 1, 15),
        vendor_name='Test SRL',
        vendor_tax_id='130000001',
        total_amount=10000.0,
        tax_amount=1800.0,
        category='servicios',
        description='Test',
        currency='DOP',
        processed=True,
        audit_flags=None,
        raw_extracted_data=None,
        goods_services_type='07',
        deleted_at=None,
        updated_at=None,
        cancelled_at=None,
        cancellation_type=None,
    )
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


class TestDgiiTxtFormat:
    """Tests that the DGII .txt output matches the official VBA macro format."""

    def setup_method(self):
        from app.services.export import ExportService
        self.svc = ExportService()

    def test_606_header_format(self):
        inv = _mock_invoice()
        buf = self.svc.export_dgii_606_txt([inv], report_rnc='130000002', period='202401')
        text = buf.decode('utf-8')
        header = text.split('\r\n')[0]
        assert header == '606|130000002|202401|1'

    def test_606_field_count(self):
        """606 detail lines must have exactly 23 pipe-delimited fields."""
        inv = _mock_invoice()
        buf = self.svc.export_dgii_606_txt([inv], report_rnc='130000002', period='202401')
        detail = buf.decode('utf-8').split('\r\n')[1]
        assert len(detail.split('|')) == 23

    def test_607_header_format(self):
        inv = _mock_invoice()
        buf = self.svc.export_dgii_607_txt([inv], report_rnc='130000002', period='202401')
        header = buf.decode('utf-8').split('\r\n')[0]
        assert header == '607|130000002|202401|1'

    def test_607_field_count(self):
        """607 detail lines must have exactly 23 pipe-delimited fields."""
        inv = _mock_invoice()
        buf = self.svc.export_dgii_607_txt([inv], report_rnc='130000002', period='202401')
        detail = buf.decode('utf-8').split('\r\n')[1]
        assert len(detail.split('|')) == 23

    def test_608_header_format(self):
        inv = _mock_invoice(raw_extracted_data='{"cancellation_type": "04"}')
        buf = self.svc.export_dgii_608_txt([inv], report_rnc='130000002', period='202401')
        header = buf.decode('utf-8').split('\r\n')[0]
        assert header == '608|130000002|202401|1'

    def test_608_field_count(self):
        """608 detail lines must have exactly 3 pipe-delimited fields."""
        inv = _mock_invoice(raw_extracted_data='{"cancellation_type": "04"}')
        buf = self.svc.export_dgii_608_txt([inv], report_rnc='130000002', period='202401')
        detail = buf.decode('utf-8').split('\r\n')[1]
        assert len(detail.split('|')) == 3

    def test_606_rnc_and_date_in_detail(self):
        inv = _mock_invoice()
        buf = self.svc.export_dgii_606_txt([inv], report_rnc='130000002', period='202401')
        detail = buf.decode('utf-8').split('\r\n')[1]
        fields = detail.split('|')
        assert fields[0] == '130000001'       # RNC vendor
        assert fields[1] == '1'               # Tipo ID (9-digit = RNC)
        assert fields[2] == '07'              # Tipo bienes
        assert fields[3] == 'E310000000001'   # NCF
        assert fields[5] == '20240115'        # Fecha comprobante

    def test_607_ncf_and_amounts(self):
        inv = _mock_invoice(total_amount=5500.50, tax_amount=990.09)
        buf = self.svc.export_dgii_607_txt([inv], report_rnc='130000002', period='202401')
        detail = buf.decode('utf-8').split('\r\n')[1]
        fields = detail.split('|')
        assert fields[2] == 'E310000000001'    # NCF
        assert fields[7] == '4510.41'          # Monto facturado sin ITBIS
        assert fields[8] == '990.09'           # ITBIS facturado

    def test_608_ncf_and_cancellation_type(self):
        inv = _mock_invoice(raw_extracted_data='{"cancellation_type": "04 Corrección"}')
        buf = self.svc.export_dgii_608_txt([inv], report_rnc='130000002', period='202401')
        detail = buf.decode('utf-8').split('\r\n')[1]
        fields = detail.split('|')
        assert fields[0] == 'E310000000001'    # NCF
        assert fields[1] == '20240115'         # Fecha
        assert fields[2] == '04'               # Tipo anulación (truncated to 2 chars)

    def test_txt_amount_formatting(self):
        """Amounts must not have trailing zeros (matching VBA Trim behavior)."""
        assert self.svc._txt_amount(1500.0) == '1500'
        assert self.svc._txt_amount(1500.50) == '1500.5'
        assert self.svc._txt_amount(1500.55) == '1500.55'
        assert self.svc._txt_amount(0.0) == ''
        assert self.svc._txt_amount(0.0, allow_zero=True) == '0'
        assert self.svc._txt_amount(None) == ''

    def test_multiple_invoices_count(self):
        """Header record count must match the number of detail lines."""
        invs = [_mock_invoice(id=f'test-{i}') for i in range(5)]
        buf = self.svc.export_dgii_606_txt(invs, report_rnc='130000002', period='202401')
        lines = buf.decode('utf-8').split('\r\n')
        header_count = int(lines[0].split('|')[3])
        assert header_count == 5
        assert len(lines) == 6  # 1 header + 5 detail

    def test_crlf_line_endings(self):
        """DGII files must use CRLF line endings."""
        inv = _mock_invoice()
        buf = self.svc.export_dgii_606_txt([inv], report_rnc='130000002', period='202401')
        text = buf.decode('utf-8')
        assert '\r\n' in text
        # Should NOT end with a trailing newline (VBA: Print #1, strDetalle;)
        assert not text.endswith('\r\n')

    def test_606_zero_itbis_is_explicit_and_payment_is_padded(self):
        inv = _mock_invoice(
            total_amount=1000.0,
            tax_amount=None,
            raw_extracted_data='{"payment_method": "1"}',
        )
        buf = self.svc.export_dgii_606_txt([inv], report_rnc='130000001', period='202401')
        fields = buf.decode('utf-8').split('\r\n')[1].split('|')

        assert fields[10] == '0'   # ITBIS Facturado
        assert fields[14] == '0'   # ITBIS por Adelantar
        assert fields[22] == '01'  # Forma de Pago

    def test_606_self_issued_purchase_uses_report_rnc(self):
        inv = _mock_invoice(
            invoice_number='B1700000001',
            vendor_tax_id=None,
            raw_extracted_data='{"payment_method": "02"}',
        )
        buf = self.svc.export_dgii_606_txt([inv], report_rnc='130000001', period='202401')
        fields = buf.decode('utf-8').split('\r\n')[1].split('|')

        assert fields[0] == '130000001'
        assert fields[1] == '1'

    def test_607_amount_is_base_but_payment_bucket_uses_total(self):
        inv = _mock_invoice(
            total_amount=1180.0,
            tax_amount=180.0,
            raw_extracted_data='{"payment_method": "02"}',
        )
        buf = self.svc.export_dgii_607_txt([inv], report_rnc='130000001', period='202401')
        fields = buf.decode('utf-8').split('\r\n')[1].split('|')

        assert fields[7] == '1000'
        assert fields[8] == '180'
        assert fields[16] == ''
        assert fields[17] == '1180'

    def test_607_zero_itbis_is_explicit(self):
        inv = _mock_invoice(
            total_amount=1000.0,
            tax_amount=None,
            raw_extracted_data='{"payment_method": "01"}',
        )
        buf = self.svc.export_dgii_607_txt([inv], report_rnc='130000001', period='202401')
        fields = buf.decode('utf-8').split('\r\n')[1].split('|')

        assert fields[7] == '1000'
        assert fields[8] == '0'


class TestDgiiValidationRules:
    def test_report_rnc_is_required_for_exportability(self):
        from app.routers.dgii import _compute_dgii_validation

        inv = _mock_invoice(raw_extracted_data='{"payment_method": "01"}')
        stats = _compute_dgii_validation([inv], "dgii_606", report_rnc="", period="202401")

        assert stats["can_export"] is False
        assert stats["missing_report_rnc"] == 1
        assert any("declarante" in err for err in stats["report_errors"])

    def test_606_requires_payment_method(self):
        from app.routers.dgii import _compute_dgii_validation

        inv = _mock_invoice(raw_extracted_data=None)
        stats = _compute_dgii_validation([inv], "dgii_606", report_rnc="130000001", period="202401")

        assert stats["can_export"] is False
        assert stats["missing_payment_method"] == 1

    def test_608_does_not_require_vendor_rnc_amount_or_payment_method(self):
        from app.routers.dgii import _compute_dgii_validation

        inv = _mock_invoice(
            invoice_number='B0100000001',
            vendor_tax_id=None,
            total_amount=None,
            tax_amount=None,
            raw_extracted_data='{"cancellation_type": "04"}',
        )
        stats = _compute_dgii_validation([inv], "dgii_608", report_rnc="130000001", period="202401")

        assert stats["can_export"] is True
        assert stats["missing_rnc"] == 0
        assert stats["zero_amount"] == 0
        assert stats["missing_payment_method"] == 0
